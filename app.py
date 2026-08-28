"""Relatório de Lojas Assaí │ Sistema de Checklist / Pontuação
Interface moderna com CustomTkinter + abas + gráfico + persistência JSON
Campos mensais: pontuação 0-100 ou texto (ADM FÉRIAS, FÉRIAS, SEM ADM)
MÉDIA = média aritmética dos valores numéricos mensais (textos excluídos)
Suporte a anexo de pontuação (visualizar/baixar)
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import sys
import re
import shutil
import subprocess
import platform
from datetime import datetime
from collections import defaultdict
import base64
import hashlib

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    import matplotlib.patheffects
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = True

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    from scipy.interpolate import make_interp_spline
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ============================================================
# CONFIGURACAO
# ============================================================
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

APP_TITLE = "Relatório de Lojas Assaí"
APP_SIZE = (1380, 850)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_DADOS = os.path.join(SCRIPT_DIR, "dados_lojas.json")
PASTA_ANEXOS = os.path.join(SCRIPT_DIR, "anexos")

VALORES_TEXTO = ["ADM FÉRIAS", "FÉRIAS", "SEM ADM"]
CAMPOS_CHAVE = ["estado", "loja", "contrato", "ano"]

MESES = [
    ("jan", "Janeiro"), ("fev", "Fevereiro"), ("mar", "Março"),
    ("abr", "Abril"), ("mai", "Maio"), ("jun", "Junho"),
    ("jul", "Julho"), ("ago", "Agosto"), ("set", "Setembro"),
    ("out", "Outubro"), ("nov", "Novembro"), ("dez", "Dezembro"),
]
MESES_CURTOS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
               "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

COLUNAS_EXPORTACAO = ["ANO", "ESTADOS", "LOJAS"] + MESES_CURTOS + ["MEDIA", "CONTRATO"]

# ============================================================
# MAPA DO BRASIL (coordenadas simplificadas dos 27 estados)
# ============================================================
MAPA_ESTADOS = {
    "AC": {
        "nome": "Acre",
        "coords": [[-73.608, -7.202], [-70.369, -8.141], [-66.627, -9.935], [-68.543, -11.109], [-70.622, -10.999], [-70.494, -9.426], [-71.212, -9.967], [-72.18, -10.0], [-72.357, -9.494], [-73.215, -9.411], [-72.937, -8.988], [-73.988, -7.555], [-73.608, -7.202]],
    },
    "AL": {
        "nome": "Alagoas",
        "coords": [[-36.952, -9.382], [-35.152, -8.913], [-35.301, -9.185], [-35.354, -9.255], [-36.391, -10.501], [-38.237, -9.329], [-37.76, -8.857], [-36.952, -9.382]],
    },
    "AM": {
        "nome": "Amazonas",
        "coords": [[-67.407, 2.247], [-67.088, 1.167], [-66.318, 0.755], [-65.585, 1.009], [-65.54, 0.649], [-65.103, 1.157], [-64.337, 1.364], [-63.996, 1.98], [-63.141, 2.173], [-62.706, 1.94], [-62.445, 0.977], [-62.188, -0.33], [-62.51, -0.759], [-61.896, -1.395], [-61.474, -1.579], [-61.585, -0.937], [-61.216, -0.5], [-60.667, -0.894], [-60.309, -0.724], [-60.037, 0.264], [-58.895, 0.264], [-58.872, -0.343], [-58.323, -1.143], [-56.679, -2.212], [-56.098, -2.027], [-58.478, -6.699], [-58.136, -7.356], [-58.395, -8.78], [-61.985, -8.878], [-62.845, -7.986], [-63.62, -7.969], [-64.149, -8.959], [-64.839, -8.994], [-65.097, -9.432], [-65.246, -9.257], [-65.791, -9.585], [-66.408, -9.407], [-66.806, -9.814], [-70.369, -8.141], [-73.804, -7.111], [-73.137, -6.497], [-73.236, -6.031], [-72.814, -5.11], [-70.653, -4.127], [-69.964, -4.3], [-69.395, -1.132], [-70.057, -0.186], [-70.043, 0.559], [-69.114, 0.65], [-69.265, 1.065], [-69.844, 1.086], [-69.842, 1.721], [-68.156, 1.732], [-68.208, 1.962], [-67.941, 1.831], [-67.407, 2.247]],
    },
    "AP": {
        "nome": "Amapá",
        "coords": [[-51.549, 4.425], [-50.702, 2.139], [-49.921, 1.704], [-49.893, 1.193], [-52.07, -1.236], [-52.933, -0.139], [-53.426, 1.243], [-54.744, 1.776], [-54.876, 2.427], [-52.944, 2.169], [-51.549, 4.425]],
    },
    "BA": {
        "nome": "Bahia",
        "coords": [[-39.288, -8.563], [-38.296, -9.022], [-37.736, -10.332], [-38.229, -10.915], [-37.813, -11.514], [-37.343, -11.443], [-38.049, -12.634], [-38.965, -13.283], [-38.856, -15.86], [-39.135, -17.688], [-39.67, -18.349], [-40.623, -17.406], [-39.856, -16.113], [-40.23, -15.803], [-41.331, -15.744], [-41.8, -15.101], [-43.176, -14.65], [-43.531, -14.815], [-44.215, -14.233], [-46.077, -15.264], [-45.907, -14.353], [-46.265, -14.098], [-46.041, -13.28], [-46.315, -13.303], [-46.114, -12.918], [-46.397, -12.04], [-46.083, -11.636], [-46.617, -11.289], [-45.603, -10.108], [-45.248, -10.822], [-44.134, -10.636], [-43.662, -10.004], [-43.849, -9.548], [-43.485, -9.265], [-42.765, -9.616], [-41.113, -8.704], [-40.623, -9.482], [-39.288, -8.563]],
    },
    "CE": {
        "nome": "Ceará",
        "coords": [[-40.018, -2.837], [-37.252, -4.832], [-37.64, -4.926], [-38.578, -6.28], [-38.765, -6.994], [-38.534, -7.293], [-38.966, -7.845], [-39.662, -7.31], [-40.548, -7.392], [-40.37, -6.803], [-40.732, -6.654], [-40.925, -5.181], [-41.249, -4.869], [-41.322, -2.921], [-40.018, -2.837]],
    },
    "ES": {
        "nome": "Espírito Santo",
        "coords": [[-40.424, -20.635], [-40.957, -21.303], [-41.718, -21.123], [-41.858, -20.372], [-41.382, -20.188], [-40.949, -19.473], [-41.158, -18.308], [-40.527, -17.891], [-39.666, -18.332], [-39.689, -19.306], [-40.424, -20.635]],
    },
    "GO": {
        "nome": "Goiás",
        "coords": [[-50.158, -12.412], [-50.292, -12.839], [-49.369, -13.274], [-49.119, -12.79], [-48.586, -13.317], [-48.173, -13.148], [-47.679, -13.467], [-47.634, -13.104], [-47.427, -13.289], [-46.114, -12.918], [-46.265, -14.098], [-45.907, -14.353], [-46.088, -14.936], [-46.503, -14.704], [-46.502, -15.052], [-46.924, -15.058], [-46.812, -15.885], [-47.319, -16.036], [-47.417, -15.5], [-48.197, -15.501], [-48.279, -16.051], [-47.304, -16.06], [-47.458, -16.502], [-47.126, -16.98], [-47.541, -17.454], [-47.283, -18.058], [-47.954, -18.5], [-48.936, -18.306], [-49.378, -18.642], [-50.309, -18.698], [-50.842, -19.499], [-52.916, -18.639], [-52.758, -18.348], [-53.101, -18.31], [-53.246, -17.532], [-52.681, -16.303], [-51.086, -14.917], [-50.871, -13.733], [-50.158, -12.412]],
    },
    "MA": {
        "nome": "Maranhão",
        "coords": [[-47.031, -8.985], [-46.466, -8.066], [-47.043, -8.053], [-47.746, -7.201], [-47.378, -6.27], [-47.5, -5.525], [-48.363, -5.168], [-48.755, -5.349], [-47.088, -3.855], [-46.151, -1.224], [-45.846, -1.045], [-45.719, -1.404], [-45.58, -1.257], [-45.41, -1.289], [-45.488, -1.431], [-45.487, -1.53], [-45.443, -1.543], [-45.448, -1.449], [-45.317, -1.318], [-45.352, -1.736], [-45.101, -1.36], [-44.816, -1.418], [-44.642, -1.624], [-44.814, -1.815], [-44.593, -1.744], [-44.326, -2.5], [-43.615, -2.219], [-41.823, -2.719], [-42.989, -4.234], [-42.919, -6.67], [-44.033, -6.76], [-45.456, -7.67], [-45.994, -8.926], [-45.946, -10.258], [-46.367, -10.168], [-47.031, -8.985]],
    },
    "MG": {
        "nome": "Minas Gerais",
        "coords": [[-44.209, -14.244], [-43.783, -14.339], [-43.883, -14.653], [-43.531, -14.815], [-43.176, -14.65], [-41.8, -15.101], [-41.331, -15.744], [-40.707, -15.666], [-39.856, -16.113], [-40.57, -17.061], [-40.222, -17.98], [-40.882, -17.97], [-40.771, -18.155], [-41.158, -18.308], [-40.944, -19.46], [-41.382, -20.188], [-41.847, -20.329], [-42.271, -21.715], [-46.345, -22.904], [-46.723, -22.306], [-46.509, -21.469], [-47.011, -21.422], [-47.466, -19.964], [-48.823, -20.161], [-48.899, -20.441], [-49.551, -19.905], [-50.471, -19.779], [-51.0, -20.085], [-50.962, -19.484], [-50.309, -18.698], [-49.378, -18.642], [-48.936, -18.306], [-47.954, -18.5], [-47.283, -18.058], [-47.538, -17.388], [-47.126, -16.98], [-47.458, -16.502], [-47.3, -16.017], [-46.812, -15.885], [-46.918, -15.049], [-46.502, -15.052], [-46.474, -14.705], [-46.003, -14.902], [-46.119, -15.192], [-46.052, -15.259], [-44.209, -14.244]],
    },
    "MS": {
        "nome": "Mato Grosso do Sul",
        "coords": [[-53.874, -17.922], [-53.071, -18.039], [-53.069, -18.342], [-52.758, -18.348], [-52.916, -18.639], [-51.057, -19.329], [-51.001, -20.096], [-52.408, -22.141], [-53.607, -22.951], [-54.129, -23.982], [-55.347, -23.994], [-55.849, -22.284], [-57.991, -22.09], [-57.819, -20.942], [-58.167, -20.171], [-57.859, -19.97], [-58.131, -19.758], [-57.453, -18.231], [-57.795, -17.56], [-57.452, -17.902], [-56.113, -17.167], [-55.127, -17.652], [-54.302, -17.661], [-53.708, -17.228], [-53.874, -17.922]],
    },
    "MT": {
        "nome": "Mato Grosso",
        "coords": [[-60.716, -13.682], [-59.774, -12.341], [-60.108, -11.839], [-59.976, -11.122], [-61.55, -10.986], [-61.582, -8.798], [-58.415, -8.792], [-58.138, -7.349], [-57.592, -8.756], [-56.761, -9.405], [-50.224, -9.841], [-50.739, -11.435], [-50.502, -12.884], [-51.086, -14.917], [-52.681, -16.303], [-53.218, -17.299], [-53.071, -18.039], [-53.948, -17.923], [-53.708, -17.228], [-54.084, -17.619], [-55.127, -17.652], [-56.113, -17.167], [-57.452, -17.902], [-58.395, -17.184], [-58.321, -16.264], [-60.171, -16.265], [-60.564, -15.108], [-60.244, -15.096], [-60.381, -13.987], [-60.716, -13.682]],
    },
    "PA": {
        "nome": "Pará",
        "coords": [[-48.472, -0.499], [-46.428, -1.065], [-46.19, -0.894], [-46.272, -1.172], [-46.072, -1.019], [-46.28, -2.153], [-47.581, -4.52], [-48.755, -5.349], [-48.138, -5.602], [-48.231, -5.946], [-49.209, -6.925], [-49.215, -8.194], [-50.224, -9.841], [-56.754, -9.406], [-57.592, -8.756], [-58.478, -6.699], [-56.098, -2.027], [-56.679, -2.212], [-58.43, -1.027], [-58.872, -0.343], [-58.895, 1.228], [-57.304, 2.0], [-55.956, 1.845], [-55.978, 2.528], [-54.954, 2.584], [-54.744, 1.776], [-53.426, 1.243], [-52.933, -0.139], [-52.1, -1.226], [-50.157, 0.705], [-50.061, 0.339], [-48.412, -0.257], [-48.472, -0.499]],
    },
    "PB": {
        "nome": "Paraíba",
        "coords": [[-37.227, -6.035], [-37.484, -6.71], [-36.718, -6.982], [-36.394, -6.294], [-34.971, -6.485], [-34.826, -7.547], [-35.479, -7.445], [-36.991, -8.303], [-37.355, -7.975], [-36.984, -7.482], [-37.233, -7.275], [-38.077, -7.83], [-38.593, -7.754], [-38.602, -6.389], [-38.115, -6.521], [-37.227, -6.035]],
    },
    "PE": {
        "nome": "Pernambuco",
        "coords": [[-37.177, -7.309], [-36.984, -7.482], [-37.355, -7.975], [-36.991, -8.303], [-35.479, -7.445], [-34.84, -7.543], [-35.152, -8.913], [-35.896, -8.854], [-36.952, -9.382], [-37.76, -8.857], [-38.237, -9.329], [-38.479, -8.85], [-39.383, -8.533], [-40.623, -9.482], [-40.921, -8.835], [-41.358, -8.707], [-40.589, -8.138], [-40.548, -7.392], [-39.662, -7.31], [-39.091, -7.858], [-38.715, -7.622], [-38.077, -7.83], [-37.177, -7.309]],
    },
    "PI": {
        "nome": "Piauí",
        "coords": [[-41.739, -2.806], [-41.257, -3.004], [-41.249, -4.869], [-40.925, -5.181], [-40.732, -6.654], [-40.37, -6.803], [-40.713, -7.473], [-40.589, -8.138], [-41.838, -9.242], [-42.765, -9.616], [-43.485, -9.265], [-43.849, -9.548], [-43.662, -10.004], [-44.13, -10.633], [-45.248, -10.822], [-45.579, -10.122], [-45.955, -10.218], [-45.994, -8.926], [-45.496, -7.75], [-44.053, -6.768], [-42.919, -6.67], [-42.989, -4.234], [-41.739, -2.806]],
    },
    "PR": {
        "nome": "Paraná",
        "coords": [[-52.972, -22.57], [-49.727, -23.108], [-49.305, -24.672], [-48.581, -24.671], [-48.556, -25.084], [-48.023, -25.23], [-48.59, -25.976], [-49.555, -26.237], [-50.571, -26.003], [-51.411, -26.717], [-53.551, -26.292], [-53.892, -25.622], [-54.593, -25.592], [-54.341, -24.129], [-53.607, -22.951], [-52.972, -22.57]],
    },
    "RJ": {
        "nome": "Rio de Janeiro",
        "coords": [[-44.668, -23.054], [-44.724, -23.368], [-44.875, -23.249], [-44.802, -22.999], [-44.161, -22.678], [-44.793, -22.387], [-42.271, -21.715], [-41.875, -20.766], [-41.718, -21.123], [-40.957, -21.303], [-40.985, -21.999], [-41.96, -22.534], [-42.013, -22.997], [-43.051, -22.982], [-43.085, -22.677], [-43.286, -23.016], [-43.553, -23.076], [-43.711, -23.056], [-43.856, -22.902], [-44.668, -23.054]],
    },
    "RN": {
        "nome": "Rio Grande do Norte",
        "coords": [[-37.64, -4.926], [-35.489, -5.158], [-34.969, -6.488], [-36.394, -6.294], [-36.718, -6.982], [-37.484, -6.71], [-37.174, -6.048], [-38.115, -6.521], [-38.577, -6.347], [-37.64, -4.926]],
    },
    "RO": {
        "nome": "Rondônia",
        "coords": [[-65.374, -9.699], [-66.806, -9.814], [-66.408, -9.407], [-65.097, -9.432], [-64.839, -8.994], [-64.143, -8.953], [-63.62, -7.969], [-62.866, -7.975], [-62.124, -8.801], [-61.468, -8.917], [-61.55, -10.986], [-59.986, -11.114], [-60.108, -11.839], [-59.774, -12.341], [-60.708, -13.692], [-61.84, -13.548], [-63.157, -12.613], [-64.291, -12.5], [-65.031, -11.995], [-65.361, -11.251], [-65.374, -9.699]],
    },
    "RR": {
        "nome": "Roraima",
        "coords": [[-59.916, 3.146], [-59.751, 1.862], [-58.886, 1.261], [-58.895, 0.264], [-60.037, 0.264], [-60.531, -0.875], [-61.087, -0.5], [-61.428, -0.634], [-61.482, -1.58], [-62.51, -0.759], [-62.188, -0.33], [-62.706, 1.94], [-64.055, 2.498], [-64.185, 3.56], [-64.824, 4.244], [-63.964, 3.868], [-63.206, 3.952], [-62.96, 3.608], [-62.747, 4.035], [-60.996, 4.518], [-60.591, 4.927], [-60.723, 5.22], [-60.21, 5.271], [-59.99, 4.987], [-60.162, 4.508], [-59.675, 4.373], [-59.517, 3.943], [-59.916, 3.146]],
    },
    "RS": {
        "nome": "Rio Grande do Sul",
        "coords": [[-51.095, -30.381], [-52.256, -31.849], [-52.098, -32.162], [-52.629, -33.116], [-53.416, -33.748], [-53.43, -33.156], [-53.123, -32.791], [-52.75, -32.862], [-52.622, -32.143], [-53.076, -32.656], [-53.388, -32.586], [-55.577, -30.833], [-56.009, -31.081], [-56.807, -30.104], [-57.643, -30.188], [-54.812, -27.529], [-53.874, -27.127], [-52.166, -27.273], [-50.625, -28.391], [-49.691, -28.618], [-50.166, -29.247], [-49.711, -29.325], [-50.769, -31.11], [-52.081, -32.157], [-52.098, -31.835], [-51.238, -31.458], [-50.574, -30.481], [-50.597, -30.194], [-50.93, -30.435], [-51.295, -30.001], [-51.095, -30.381]],
    },
    "SC": {
        "nome": "Santa Catarina",
        "coords": [[-53.834, -27.169], [-53.643, -26.252], [-53.281, -26.247], [-51.411, -26.717], [-50.571, -26.003], [-49.555, -26.237], [-48.643, -25.956], [-48.465, -27.145], [-48.616, -27.251], [-48.743, -28.508], [-50.065, -29.341], [-49.765, -28.46], [-50.625, -28.391], [-52.166, -27.273], [-53.834, -27.169]],
    },
    "SE": {
        "nome": "Sergipe",
        "coords": [[-37.96, -9.533], [-36.393, -10.498], [-37.518, -11.548], [-38.24, -10.876], [-37.736, -10.332], [-37.96, -9.533]],
    },
    "SP": {
        "nome": "São Paulo",
        "coords": [[-46.138, -23.859], [-48.099, -25.311], [-48.251, -24.978], [-48.582, -25.051], [-48.581, -24.671], [-49.305, -24.672], [-49.2, -24.344], [-49.986, -22.897], [-53.108, -22.597], [-52.408, -22.141], [-51.586, -20.634], [-50.576, -19.816], [-49.265, -19.962], [-48.899, -20.441], [-48.823, -20.161], [-47.473, -19.962], [-47.011, -21.422], [-46.509, -21.469], [-46.723, -22.306], [-46.359, -22.896], [-44.809, -22.405], [-44.162, -22.674], [-44.792, -22.982], [-44.724, -23.368], [-45.406, -23.623], [-45.405, -23.82], [-46.138, -23.859]],
    },
    "TO": {
        "nome": "Tocantins",
        "coords": [[-47.033, -8.982], [-46.367, -10.168], [-45.699, -10.166], [-46.617, -11.289], [-46.086, -11.622], [-46.397, -12.04], [-46.119, -12.925], [-47.427, -13.289], [-47.634, -13.104], [-47.679, -13.467], [-48.173, -13.148], [-48.586, -13.317], [-48.857, -12.805], [-49.369, -13.274], [-50.292, -12.839], [-50.142, -12.396], [-50.622, -12.819], [-50.603, -10.661], [-49.283, -8.379], [-49.209, -6.925], [-48.382, -6.379], [-48.132, -5.618], [-48.745, -5.369], [-48.519, -5.192], [-47.5, -5.525], [-47.378, -6.27], [-47.746, -7.201], [-47.043, -8.053], [-46.477, -8.012], [-47.033, -8.982]],
    },
    "DF": {
        "nome": "Distrito Federal",
        "coords": [[-47.417, -15.5], [-47.308, -16.05], [-48.279, -16.051], [-48.197, -15.501], [-47.417, -15.5]],
    },
}

ESTADO_PARA_SIGLA = {
    "CEARA": "CE",
    "BAHIA": "BA",
    "SERGIPE": "SE",
    "PERNAMBUCO": "PE",
    "PARAIBA": "PB",
    "RIO GRANDE DO NORTE": "RN",
    "PIAUI": "PI",
    "MARANHAO": "MA",
    "AMAPA": "AP",
    "MANAUS": "AM",
    "PARA": "PA",
    "RORAIMA": "RR",
}


# Criar pasta de anexos se nao existir
os.makedirs(PASTA_ANEXOS, exist_ok=True)


# ============================================================
# DADOS REAIS (148 registros, 12 estados, 39 lojas, 2023-2026)
# ============================================================
def _reg(estado, loja, contrato, ano):
    """Cria registro base com campos mensais de pontuacao (string)."""
    reg = {"estado": estado, "loja": loja, "contrato": contrato, "ano": ano}
    for mk, _ in MESES:
        reg[f"pont_{mk}"] = ""
    reg["media"] = 0.0
    reg["anexo"] = ""  # caminho do anexo
    return reg


DADOS_REAIS = [
    # CEARA (8 lojas x 4 anos = 32)
    _reg("CEARA","ASSAÍ - CAUCAIA","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - CAUCAIA","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - CAUCAIA","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - CAUCAIA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - PARANGABA","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - PARANGABA","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - PARANGABA","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - PARANGABA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MESSEJANA","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - MESSEJANA","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - MESSEJANA","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - MESSEJANA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MISTER HALL","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - MISTER HALL","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - MISTER HALL","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - MISTER HALL","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - SOBRAL","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - SOBRAL","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - SOBRAL","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - SOBRAL","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - WASHIGUNTON SOARES","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - WASHIGUNTON SOARES","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - WASHIGUNTON SOARES","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - WASHIGUNTON SOARES","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - BEZERRA DE MENEZES","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - BEZERRA DE MENEZES","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - BEZERRA DE MENEZES","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - BEZERRA DE MENEZES","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MARACANAÚ","ASSAÍ - ATACADISTA",2023),
    _reg("CEARA","ASSAÍ - MARACANAÚ","ASSAÍ - ATACADISTA",2024),
    _reg("CEARA","ASSAÍ - MARACANAÚ","ASSAÍ - ATACADISTA",2025),
    _reg("CEARA","ASSAÍ - MARACANAÚ","ASSAÍ - ATACADISTA",2026),

    # BAHIA (6 lojas x 4 anos = 24)
    _reg("BAHIA","ASSAÍ - CAMASSARI","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - CAMASSARI","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - CAMASSARI","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - CAMASSARI","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - FEIRA DE SANTANA","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - FEIRA DE SANTANA","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - FEIRA DE SANTANA","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - FEIRA DE SANTANA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - GUANAMBÍ","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - GUANAMBÍ","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - GUANAMBÍ","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - GUANAMBÍ","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - MUSSURUNGA","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - MUSSURUNGA","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - MUSSURUNGA","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - MUSSURUNGA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - VITÓRIA DA CONQUISTA","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - VITÓRIA DA CONQUISTA","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - VITÓRIA DA CONQUISTA","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - VITÓRIA DA CONQUISTA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - JEQUIÉ","ASSAÍ - ATACADISTA",2023),
    _reg("BAHIA","ASSAÍ - JEQUIÉ","ASSAÍ - ATACADISTA",2024),
    _reg("BAHIA","ASSAÍ - JEQUIÉ","ASSAÍ - ATACADISTA",2025),
    _reg("BAHIA","ASSAÍ - JEQUIÉ","ASSAÍ - ATACADISTA",2026),

    # PERNAMBUCO (5 lojas x 4 anos = 20)
    _reg("PERNAMBUCO","ASSAÍ - IMBIRIBEIRA","ASSAÍ - ATACADISTA",2023),
    _reg("PERNAMBUCO","ASSAÍ - IMBIRIBEIRA","ASSAÍ - ATACADISTA",2024),
    _reg("PERNAMBUCO","ASSAÍ - IMBIRIBEIRA","ASSAÍ - ATACADISTA",2025),
    _reg("PERNAMBUCO","ASSAÍ - IMBIRIBEIRA","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - AV.RECIFE","ASSAÍ - ATACADISTA",2023),
    _reg("PERNAMBUCO","ASSAÍ - AV.RECIFE","ASSAÍ - ATACADISTA",2024),
    _reg("PERNAMBUCO","ASSAÍ - AV.RECIFE","ASSAÍ - ATACADISTA",2025),
    _reg("PERNAMBUCO","ASSAÍ - AV.RECIFE","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - CAMARAGIBE","ASSAÍ - ATACADISTA",2023),
    _reg("PERNAMBUCO","ASSAÍ - CAMARAGIBE","ASSAÍ - ATACADISTA",2024),
    _reg("PERNAMBUCO","ASSAÍ - CAMARAGIBE","ASSAÍ - ATACADISTA",2025),
    _reg("PERNAMBUCO","ASSAÍ - CAMARAGIBE","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - CARUARU","ASSAÍ - ATACADISTA",2023),
    _reg("PERNAMBUCO","ASSAÍ - CARUARU","ASSAÍ - ATACADISTA",2024),
    _reg("PERNAMBUCO","ASSAÍ - CARUARU","ASSAÍ - ATACADISTA",2025),
    _reg("PERNAMBUCO","ASSAÍ - CARUARU","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - PETROLINA","ASSAÍ - ATACADISTA",2023),
    _reg("PERNAMBUCO","ASSAÍ - PETROLINA","ASSAÍ - ATACADISTA",2024),
    _reg("PERNAMBUCO","ASSAÍ - PETROLINA","ASSAÍ - ATACADISTA",2025),
    _reg("PERNAMBUCO","ASSAÍ - PETROLINA","ASSAÍ - ATACADISTA",2026),

    # SERGIPE (3 lojas x 4 anos = 12)
    _reg("SERGIPE","ASSAÍ - ARACAJU","ASSAÍ - ATACADISTA",2023),
    _reg("SERGIPE","ASSAÍ - ARACAJU","ASSAÍ - ATACADISTA",2024),
    _reg("SERGIPE","ASSAÍ - ARACAJU","ASSAÍ - ATACADISTA",2025),
    _reg("SERGIPE","ASSAÍ - ARACAJU","ASSAÍ - ATACADISTA",2026),
    _reg("SERGIPE","ASSAÍ - LAGARTO","ASSAÍ - ATACADISTA",2023),
    _reg("SERGIPE","ASSAÍ - LAGARTO","ASSAÍ - ATACADISTA",2024),
    _reg("SERGIPE","ASSAÍ - LAGARTO","ASSAÍ - ATACADISTA",2025),
    _reg("SERGIPE","ASSAÍ - LAGARTO","ASSAÍ - ATACADISTA",2026),
    _reg("SERGIPE","ASSAÍ - ITABAIANA","ASSAÍ - ATACADISTA",2023),
    _reg("SERGIPE","ASSAÍ - ITABAIANA","ASSAÍ - ATACADISTA",2024),
    _reg("SERGIPE","ASSAÍ - ITABAIANA","ASSAÍ - ATACADISTA",2025),
    _reg("SERGIPE","ASSAÍ - ITABAIANA","ASSAÍ - ATACADISTA",2026),

    # PARA (5 lojas x 4 anos = 20, com variacoes)
    _reg("PARA","ASSAÍ - BELÉM","ASSAÍ - ATACADISTA",2023),
    _reg("PARA","ASSAÍ - BELÉM","ASSAÍ - ATACADISTA",2024),
    _reg("PARA","ASSAÍ - BELÉM","ASSAÍ - ATACADISTA",2025),
    _reg("PARA","ASSAÍ - BELÉM","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","BELÉM - PA","ASSAÍ - ATACADISTA",2023),
    _reg("PARA","ASSAÍ - ANANINDEUA","ASSAÍ - ATACADISTA",2023),
    _reg("PARA","ASSAÍ - ANANINDEUA","ASSAÍ - ATACADISTA",2024),
    _reg("PARA","ASSAÍ - ANANINDEUA","ASSAÍ - ATACADISTA",2025),
    _reg("PARA","ASSAÍ - ANANINDEUA","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - CASTANHAL","ASSAÍ - ATACADISTA",2023),
    _reg("PARA","ASSAÍ - CASTANHAL","ASSAÍ - ATACADISTA",2024),
    _reg("PARA","ASSAÍ - CASTANHAL","ASSAÍ - ATACADISTA",2025),
    _reg("PARA","ASSAÍ - CASTANHAL","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - MARABÁ","ASSAÍ - ATACADISTA",2023),
    _reg("PARA","ASSAÍ - MARABÁ","ASSAÍ - ATACADISTA",2024),
    _reg("PARA","ASSAÍ - MARABÁ","ASSAÍ - ATACADISTA",2025),
    _reg("PARA","ASSAÍ - MARABÁ","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ALGUSTO MONTE NEGRO","ASSAÍ - ATACADISTA",2025),
    _reg("PARA","ALGUSTO MONTE NEGRO","ASSAÍ - ATACADISTA",2026),

    # MARANHAO (3 lojas x 4 anos = 12)
    _reg("MARANHAO","ASSAÍ - SÃO LUIS","ASSAÍ - ATACADISTA",2023),
    _reg("MARANHAO","ASSAÍ - SÃO LUIS","ASSAÍ - ATACADISTA",2024),
    _reg("MARANHAO","ASSAÍ - SÃO LUIS","ASSAÍ - ATACADISTA",2025),
    _reg("MARANHAO","ASSAÍ - SÃO LUIS","ASSAÍ - ATACADISTA",2026),
    _reg("MARANHAO","ASSAÍ - IMPERATRIZ","ASSAÍ - ATACADISTA",2023),
    _reg("MARANHAO","ASSAÍ - IMPERATRIZ","ASSAÍ - ATACADISTA",2024),
    _reg("MARANHAO","ASSAÍ - IMPERATRIZ","ASSAÍ - ATACADISTA",2025),
    _reg("MARANHAO","ASSAÍ - IMPERATRIZ","ASSAÍ - ATACADISTA",2026),
    _reg("MARANHAO","ASSAÍ - TIMON","ASSAÍ - ATACADISTA",2023),
    _reg("MARANHAO","ASSAÍ - TIMON","ASSAÍ - ATACADISTA",2024),
    _reg("MARANHAO","ASSAÍ - TIMON","ASSAÍ - ATACADISTA",2025),
    _reg("MARANHAO","ASSAÍ - TIMON","ASSAÍ - ATACADISTA",2026),

    # PIAUI (3 lojas x 4 anos = 12)
    _reg("PIAUI","ASSAÍ - TERESINA","ASSAÍ - ATACADISTA",2023),
    _reg("PIAUI","ASSAÍ - TERESINA","ASSAÍ - ATACADISTA",2024),
    _reg("PIAUI","ASSAÍ - TERESINA","ASSAÍ - ATACADISTA",2025),
    _reg("PIAUI","ASSAÍ - TERESINA","ASSAÍ - ATACADISTA",2026),
    _reg("PIAUI","ASSAÍ - PIRIPIRI","ASSAÍ - ATACADISTA",2023),
    _reg("PIAUI","ASSAÍ - PIRIPIRI","ASSAÍ - ATACADISTA",2024),
    _reg("PIAUI","ASSAÍ - PIRIPIRI","ASSAÍ - ATACADISTA",2025),
    _reg("PIAUI","ASSAÍ - PIRIPIRI","ASSAÍ - ATACADISTA",2026),
    _reg("PIAUI","ASSAÍ - PICOS","ASSAÍ - ATACADISTA",2023),
    _reg("PIAUI","ASSAÍ - PICOS","ASSAÍ - ATACADISTA",2024),
    _reg("PIAUI","ASSAÍ - PICOS","ASSAÍ - ATACADISTA",2025),
    _reg("PIAUI","ASSAÍ - PICOS","ASSAÍ - ATACADISTA",2026),

    # RIO GRANDE DO NORTE (4 lojas x 4 anos = 16)
    _reg("RIO GRANDE DO NORTE","ASSAÍ - NATAL","ASSAÍ - ATACADISTA",2023),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - NATAL","ASSAÍ - ATACADISTA",2024),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - NATAL","ASSAÍ - ATACADISTA",2025),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - NATAL","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - ARAPIRACA","ASSAÍ - ATACADISTA",2023),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - ARAPIRACA","ASSAÍ - ATACADISTA",2024),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - ARAPIRACA","ASSAÍ - ATACADISTA",2025),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - ARAPIRACA","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - MOSSORÓ","ASSAÍ - ATACADISTA",2023),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - MOSSORÓ","ASSAÍ - ATACADISTA",2024),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - MOSSORÓ","ASSAÍ - ATACADISTA",2025),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - MOSSORÓ","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - PARNAMIRIM","ASSAÍ - ATACADISTA",2023),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - PARNAMIRIM","ASSAÍ - ATACADISTA",2024),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - PARNAMIRIM","ASSAÍ - ATACADISTA",2025),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - PARNAMIRIM","ASSAÍ - ATACADISTA",2026),

    # PARAIBA (3 lojas x 4 anos = 12)
    _reg("PARAIBA","ASSAÍ - JOÃO PESSOA","ASSAÍ - ATACADISTA",2023),
    _reg("PARAIBA","ASSAÍ - JOÃO PESSOA","ASSAÍ - ATACADISTA",2024),
    _reg("PARAIBA","ASSAÍ - JOÃO PESSOA","ASSAÍ - ATACADISTA",2025),
    _reg("PARAIBA","ASSAÍ - JOÃO PESSOA","ASSAÍ - ATACADISTA",2026),
    _reg("PARAIBA","ASSAÍ - CAMPINA GRANDE","ASSAÍ - ATACADISTA",2023),
    _reg("PARAIBA","ASSAÍ - CAMPINA GRANDE","ASSAÍ - ATACADISTA",2024),
    _reg("PARAIBA","ASSAÍ - CAMPINA GRANDE","ASSAÍ - ATACADISTA",2025),
    _reg("PARAIBA","ASSAÍ - CAMPINA GRANDE","ASSAÍ - ATACADISTA",2026),
    _reg("PARAIBA","ASSAÍ - SANTA RITA","ASSAÍ - ATACADISTA",2023),
    _reg("PARAIBA","ASSAÍ - SANTA RITA","ASSAÍ - ATACADISTA",2024),
    _reg("PARAIBA","ASSAÍ - SANTA RITA","ASSAÍ - ATACADISTA",2025),
    _reg("PARAIBA","ASSAÍ - SANTA RITA","ASSAÍ - ATACADISTA",2026),

    # MANAUS (2 lojas x 3 anos = 6, comeca 2024)
    _reg("MANAUS","ASSAÍ - MANAUS","ASSAÍ - ATACADISTA",2024),
    _reg("MANAUS","ASSAÍ - MANAUS","ASSAÍ - ATACADISTA",2025),
    _reg("MANAUS","ASSAÍ - MANAUS","ASSAÍ - ATACADISTA",2026),
    _reg("MANAUS","ASSAÍ - MANAUS II","ASSAÍ - ATACADISTA",2024),
    _reg("MANAUS","ASSAÍ - MANAUS II","ASSAÍ - ATACADISTA",2025),
    _reg("MANAUS","ASSAÍ - MANAUS II","ASSAÍ - ATACADISTA",2026),

    # AMAPA (1 loja x 3 anos = 3, comeca 2024)
    _reg("AMAPA","ASSAÍ - MACAPÁ","ASSAÍ - ATACADISTA",2024),
    _reg("AMAPA","ASSAÍ - MACAPÁ","ASSAÍ - ATACADISTA",2025),
    _reg("AMAPA","ASSAÍ - MACAPÁ","ASSAÍ - ATACADISTA",2026),

    # RORAIMA (1 loja x 4 anos = 4)
    _reg("RORAIMA","ASSAÍ - BOA VISTA","ASSAÍ - ATACADISTA",2023),
    _reg("RORAIMA","ASSAÍ - BOA VISTA","ASSAÍ - ATACADISTA",2024),
    _reg("RORAIMA","ASSAÍ - BOA VISTA","ASSAÍ - ATACADISTA",2025),
    _reg("RORAIMA","ASSAÍ - BOA VISTA","ASSAÍ - ATACADISTA",2026),
]


# ============================================================
# FUNCOES UTILITARIAS
# ============================================================
def safe_float(v, default=0.0):
    """Converte valor para float, aceitando virgula como decimal."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return default


def eh_valor_texto(val):
    """Verifica se o valor e um texto de status (nao numerico)."""
    if not isinstance(val, str):
        val = str(val).strip()
    return val.strip().upper() in [v.upper() for v in VALORES_TEXTO]


def eh_valor_numerico(val):
    """Verifica se o valor e numerico (pode ser convertido para float)."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return False
    try:
        f = float(str(val).strip().replace(",", "."))
        return True
    except (ValueError, TypeError):
        return False


def calcular_media(registro):
    """Calcula a MEDIA = media aritmetica dos valores numericos mensais.
    Valores de texto (ADM FERIAS, FERIAS, SEM ADM) sao excluidos.
    """
    valores = []
    for mk, _ in MESES:
        val = str(registro.get(f"pont_{mk}", "")).strip()
        if val and not eh_valor_texto(val):
            try:
                f = float(val.replace(",", "."))
                valores.append(f)
            except (ValueError, TypeError):
                pass
    if valores:
        registro["media"] = round(sum(valores) / len(valores), 2)
    else:
        registro["media"] = 0.0
    return registro


def migrar_registro(reg):
    """Garante que o registro tem todos os campos de pontuacao mensal e anexo."""
    for mk, _ in MESES:
        campo = f"pont_{mk}"
        if campo not in reg:
            reg[campo] = ""
    if "media" not in reg:
        calcular_media(reg)
    if "anexo" not in reg:
        reg["anexo"] = ""
    return reg


def formatar_pontuacao(val):
    """Formata pontuacao para exibicao preservando a precisao exata.
    Ex: 84 -> '84,00', 84.9 -> '84,90', 84.92 -> '84,92'.
    Sempre mostra 2 casas decimais.
    """
    try:
        f = float(str(val).replace(",", "."))
        txt = f"{f:.2f}".replace(".", ",")
        return txt
    except (ValueError, TypeError):
        return str(val)


def formatar_media(val):
    """Formata MEDIA para exibicao — sempre 2 casas decimais."""
    try:
        f = float(str(val).replace(",", "."))
        txt = f"{f:.2f}".replace(".", ",")
        return txt
    except (ValueError, TypeError):
        return ""


def cor_por_pontuacao(pont):
    """Retorna cor baseada na pontuacao: verde >= 90, amarelo >= 70, vermelho < 70."""
    try:
        val = float(str(pont).replace(",", "."))
        if val >= 90:
            return "#16a34a"
        elif val >= 70:
            return "#d97706"
        else:
            return "#dc2626"
    except (ValueError, TypeError):
        return "#6b7280"


def tag_por_pontuacao(pont):
    try:
        val = float(str(pont).replace(",", "."))
        if val >= 90:
            return "excelente"
        elif val >= 70:
            return "bom"
        elif val > 0:
            return "baixo"
        else:
            return "vazio"
    except (ValueError, TypeError):
        return "texto"


def obter_estados(dados):
    vistos = []
    for r in dados:
        e = str(r.get("estado", "")).strip()
        if e and e not in vistos:
            vistos.append(e)
    return vistos


def obter_lojas_por_estado(dados, estado):
    vistos = []
    for r in dados:
        if str(r.get("estado", "")).strip().upper() == estado.upper():
            l = str(r.get("loja", "")).strip()
            if l and l not in vistos:
                vistos.append(l)
    return vistos


def obter_anos(dados):
    anos = set()
    for r in dados:
        try:
            a = int(r.get("ano", 0))
            if a > 0:
                anos.add(a)
        except (ValueError, TypeError):
            pass
    return sorted(anos)


def obter_anos_por_estado_loja(dados, estado, loja=None):
    """Retorna anos disponiveis filtrados por estado (e opcionalmente por loja)."""
    anos = set()
    for r in dados:
        if str(r.get("estado", "")).strip().upper() != estado.upper():
            continue
        if loja and loja != "TODAS AS LOJAS" and str(r.get("loja", "")).strip().upper() != loja.upper():
            continue
        try:
            a = int(r.get("ano", 0))
            if a > 0:
                anos.add(a)
        except (ValueError, TypeError):
            pass
    return sorted(anos)


def _chave_registro(r):
    """Chave unica: estado|loja|ano"""
    return (str(r.get("estado", "")).strip().upper(),
            str(r.get("loja", "")).strip().upper(),
            int(r.get("ano", 0)))


def _nome_arquivo_anexo(estado, loja, ano):
    """Gera nome de arquivo seguro para anexo."""
    safe = re.sub(r'[^\w\s\-\.]', '', f"{estado}_{loja}_{ano}")
    safe = re.sub(r'\s+', '_', safe.strip())
    return safe


def _abrir_arquivo(caminho):
    """Abre arquivo com aplicativo padrao do sistema."""
    try:
        if platform.system() == "Windows":
            os.startfile(caminho)
        elif platform.system() == "Darwin":
            subprocess.call(["open", caminho])
        else:
            subprocess.call(["xdg-open", caminho])
    except Exception:
        pass


# ============================================================
# PERSISTENCIA JSON
# ============================================================
def _migrar_dados_antigos(dados):
    """Converte dados do formato antigo (meta/realizado) para pontuacao."""
    convertidos = 0
    novos_dados = []
    for r in dados:
        novo = dict(r)
        # Se tem campos antigos, converter
        if any(f"meta_{mk}" in r or f"realizado_{mk}" in r for mk, _ in MESES):
            for mk, _ in MESES:
                campo = f"pont_{mk}"
                if campo not in novo:
                    meta_v = safe_float(r.get(f"meta_{mk}", 0))
                    real_v = safe_float(r.get(f"realizado_{mk}", 0))
                    if meta_v > 0 and real_v > 0:
                        pct = (real_v / meta_v) * 100
                        novo[campo] = formatar_pontuacao(pct)
                    elif meta_v > 0:
                        novo[campo] = "0"
                    else:
                        novo[campo] = ""
                # Remover campos antigos
                novo.pop(f"meta_{mk}", None)
                novo.pop(f"realizado_{mk}", None)
            convertidos += 1
        # Garantir campo anexo
        if "anexo" not in novo:
            novo["anexo"] = ""
        calcular_media(novo)
        _normalizar_pontuacoes(novo)
        novos_dados.append(novo)
    if convertidos > 0:
        print(f"Migrados {convertidos} registros do formato antigo para pontuacao.")
    return novos_dados


def _normalizar_pontuacoes(reg):
    """Garante que todos os campos pont_X do registro sao strings formatadas.
    Valores numericos (float/int) no JSON sao convertidos para string
    com 2 casas decimais, preservando a precisao exata.
    Ex: 84.9 (float) -> '84,90', '84,92' (str) mantido como esta.
    """
    for mk, _ in MESES:
        campo = f"pont_{mk}"
        val = reg.get(campo, "")
        if val is None or val == "":
            reg[campo] = ""
        elif isinstance(val, (int, float)):
            # Valor numerico vindo do JSON — converter para string formatada
            if eh_valor_texto(str(val)):
                reg[campo] = str(val)
            else:
                try:
                    f = float(val)
                    if 0 <= f <= 100:
                        reg[campo] = formatar_pontuacao(f)
                    else:
                        reg[campo] = formatar_pontuacao(f)
                except (ValueError, TypeError):
                    reg[campo] = str(val)
        elif isinstance(val, str):
            val_s = val.strip()
            if not val_s:
                reg[campo] = ""
            elif eh_valor_texto(val_s):
                reg[campo] = val_s  # texto de status preservado
            else:
                try:
                    f = float(val_s.replace(",", "."))
                    reg[campo] = formatar_pontuacao(f)
                except (ValueError, TypeError):
                    reg[campo] = val_s
    # Normalizar media tambem
    media_val = reg.get("media", 0)
    if isinstance(media_val, (int, float)):
        reg["media"] = round(float(media_val), 2)
    elif isinstance(media_val, str) and media_val.strip():
        try:
            reg["media"] = round(float(media_val.replace(",", ".")), 2)
        except (ValueError, TypeError):
            reg["media"] = 0.0
    return reg


def carregar_dados():
    """Carrega dados do JSON ou gera dados padrao. Auto-detecta corrupcao."""
    if os.path.exists(ARQUIVO_DADOS):
        try:
            with open(ARQUIVO_DADOS, "r", encoding="utf-8") as f:
                conteudo = f.read().strip()
            if not conteudo:
                raise ValueError("Arquivo vazio")
            dados = json.loads(conteudo)
            if not isinstance(dados, list):
                raise ValueError("Formato invalido")
            # Migrar dados antigos se necessario
            dados = _migrar_dados_antigos(dados)
            # Garantir campos em cada registro
            for r in dados:
                migrar_registro(r)
                # Normalizar pontuacoes: garantir que sao strings formatadas
                _normalizar_pontuacoes(r)
            return dados
        except (json.JSONDecodeError, ValueError) as e:
            print(f"JSON corrompido ou invalido: {e}")
            print("Restaurando dados padrao...")
            try:
                # Backup do arquivo corrompido
                bak = ARQUIVO_DADOS + ".corrompido_" + datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copy2(ARQUIVO_DADOS, bak)
                os.remove(ARQUIVO_DADOS)
            except Exception:
                pass
            dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(dados)
            return dados
        except Exception as e:
            print(f"Erro inesperado ao carregar: {e}")
            try:
                os.remove(ARQUIVO_DADOS)
            except Exception:
                pass
            dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(dados)
            return dados

    # Primeira execucao: gerar dados padrao
    dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
    salvar_dados_json(dados)
    return dados


def salvar_dados_json(dados):
    """Salva dados em JSON com formatacao legivel."""
    try:
        with open(ARQUIVO_DADOS, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Erro ao salvar JSON: {e}")


# ============================================================
# APLICACAO PRINCIPAL
# ============================================================
class RelatorioLojasApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry(f"{APP_SIZE[0]}x{APP_SIZE[1]}")
        self.minsize(1100, 700)

        try:
            self.dados = carregar_dados()
        except Exception:
            self.dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            try:
                salvar_dados_json(self.dados)
            except Exception:
                pass

        self._debounce_id = None
        self._atualizando_ui = False
        self._janela_anexo = None

        self._criar_toolbar()
        self._criar_abas()
        self._popular_combos()
        self._atualizar_resumo()

    # ---- TOOLBAR ----
    def _criar_toolbar(self):
        toolbar = ctk.CTkFrame(self, height=50, fg_color="#2d2d2d", corner_radius=0)
        toolbar.pack(fill="x", side="top")
        toolbar.pack_propagate(False)

        titulo = ctk.CTkLabel(toolbar, text="Checklist de Pontuação │ Lojas Assaí",
                              font=ctk.CTkFont(size=18, weight="bold"), text_color="white")
        titulo.pack(side="left", padx=15)

        self.lbl_contador = ctk.CTkLabel(toolbar, text="", font=ctk.CTkFont(size=12),
                                          text_color="#a0a0a0")
        self.lbl_contador.pack(side="left", padx=15)

        btn_export = ctk.CTkButton(toolbar, text="Exportar Excel", fg_color="#3b82f6",
                                    hover_color="#2563eb", width=150, command=self._exportar_excel)
        btn_export.pack(side="right", padx=10, pady=8)

        btn_reset = ctk.CTkButton(toolbar, text="Resetar Dados", fg_color="#f59e0b",
                                    hover_color="#d97706", width=150, command=self._resetar_dados)
        btn_reset.pack(side="right", padx=5, pady=8)

        if HAS_OPENPYXL:
            btn_import = ctk.CTkButton(toolbar, text="Importar Excel", fg_color="#22c55e",
                                        hover_color="#16a34a", width=160, command=self._importar_excel)
            btn_import.pack(side="right", padx=5, pady=8)

    # ---- ABAS ----
    def _criar_abas(self):
        self.tabview = ctk.CTkTabview(self, fg_color="#f0f0f0")
        self.tabview.pack(fill="both", expand=True, padx=10, pady=5)

        self.tab_pesquisa = self.tabview.add("Pesquisa e Edicao")
        self.tab_tabela = self.tabview.add("Tabela de Dados")

        self._criar_aba_pesquisa()
        self._criar_aba_tabela()

    # ---- ABA PESQUISA ----
    def _criar_aba_pesquisa(self):
        # Painel esquerdo: filtros + edicao (scrollavel)
        painel_esq = ctk.CTkFrame(self.tab_pesquisa, width=420, corner_radius=10)
        painel_esq.pack(side="left", fill="both", padx=(10, 5), pady=10)
        painel_esq.pack_propagate(False)

        # Painel direito: grafico
        painel_dir = ctk.CTkFrame(self.tab_pesquisa, corner_radius=10)
        painel_dir.pack(side="right", fill="both", expand=True, padx=(5, 10), pady=10)

        # ---- FILTROS ----
        frm_filtros = ctk.CTkFrame(painel_esq, fg_color="transparent")
        frm_filtros.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(frm_filtros, text="Estado:", font=ctk.CTkFont(size=12)).grid(row=0, column=0, sticky="w", padx=5)
        self.combo_estado = ctk.CTkComboBox(frm_filtros, values=[], width=200,
                                             command=self._ao_mudar_estado)
        self.combo_estado.grid(row=0, column=1, padx=5, pady=3)
        ctk.CTkButton(frm_filtros, text="➕", width=32, height=28,
                      command=self._novo_estado).grid(row=0, column=2, padx=(2, 5), pady=3)

        ctk.CTkLabel(frm_filtros, text="Loja:", font=ctk.CTkFont(size=12)).grid(row=1, column=0, sticky="w", padx=5)
        self.combo_loja = ctk.CTkComboBox(frm_filtros, values=[], width=270,
                                            command=self._ao_mudar_loja)
        self.combo_loja.grid(row=1, column=1, padx=5, pady=3)
        ctk.CTkButton(frm_filtros, text="➕", width=32, height=28,
                      command=self._nova_loja).grid(row=1, column=2, padx=(2, 5), pady=3)

        ctk.CTkLabel(frm_filtros, text="Ano:", font=ctk.CTkFont(size=12)).grid(row=2, column=0, sticky="w", padx=5)
        self.combo_ano = ctk.CTkComboBox(frm_filtros, values=[], width=120,
                                          command=self._ao_mudar_ano)
        self.combo_ano.grid(row=2, column=1, sticky="w", padx=5, pady=3)
        ctk.CTkButton(frm_filtros, text="➕", width=32, height=28,
                      command=self._novo_ano).grid(row=2, column=2, padx=(2, 5), pady=3)

        # Contrato
        ctk.CTkLabel(frm_filtros, text="Contrato:", font=ctk.CTkFont(size=12)).grid(row=3, column=0, sticky="w", padx=5)
        self.entry_contrato = ctk.CTkEntry(frm_filtros, width=270, placeholder_text="Contrato")
        self.entry_contrato.grid(row=3, column=1, padx=5, pady=3)
        self.entry_contrato.insert(0, "ASSAÍ - ATACADISTA")

        # (Mapa do Brasil movido para fundo do grafico — ver _desenhar_mapa_fundo)

        # ---- SECAO DE ANEXO ----
        frm_anexo = ctk.CTkFrame(painel_esq, fg_color="#f0f4ff", corner_radius=8, border_width=1, border_color="#c7d2fe")
        frm_anexo.pack(fill="x", padx=10, pady=3)

        ctk.CTkLabel(frm_anexo, text="Anexo da Pontuacao",
                     font=ctk.CTkFont(size=13, weight="bold"), text_color="#4338ca").pack(anchor="w", padx=10, pady=(8, 2))

        frm_anexo_btns = ctk.CTkFrame(frm_anexo, fg_color="transparent")
        frm_anexo_btns.pack(fill="x", padx=10, pady=3)

        self.btn_anexar = ctk.CTkButton(frm_anexo_btns, text="Anexar", fg_color="#6366f1",
                                         hover_color="#4f46e5", width=100, command=self._anexar_arquivo)
        self.btn_anexar.pack(side="left", padx=3)

        self.btn_visualizar = ctk.CTkButton(frm_anexo_btns, text="Visualizar", fg_color="#8b5cf6",
                                             hover_color="#7c3aed", width=100, command=self._visualizar_anexo,
                                             state="disabled")
        self.btn_visualizar.pack(side="left", padx=3)

        self.btn_baixar = ctk.CTkButton(frm_anexo_btns, text="Baixar", fg_color="#8b5cf6",
                                         hover_color="#7c3aed", width=100, command=self._baixar_anexo,
                                         state="disabled")
        self.btn_baixar.pack(side="left", padx=3)

        self.btn_remover_anexo = ctk.CTkButton(frm_anexo_btns, text="Remover", fg_color="#ef4444",
                                                hover_color="#dc2626", width=40, command=self._remover_anexo,
                                                state="disabled")
        self.btn_remover_anexo.pack(side="left", padx=3)

        self.lbl_anexo_status = ctk.CTkLabel(frm_anexo, text="Nenhum anexo",
                                              font=ctk.CTkFont(size=11), text_color="#6b7280")
        self.lbl_anexo_status.pack(anchor="w", padx=10, pady=(0, 8))

        # ---- MÉDIA ----
        frm_media = ctk.CTkFrame(painel_esq, fg_color="transparent")
        frm_media.pack(fill="x", padx=10, pady=3)
        ctk.CTkLabel(frm_media, text="MÉDIA:", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        self.lbl_media = ctk.CTkLabel(frm_media, text="", font=ctk.CTkFont(size=22, weight="bold"),
                                        text_color="#6b7280")
        self.lbl_media.pack(side="left", padx=10)
        self.lbl_status_media = ctk.CTkLabel(frm_media, text="", font=ctk.CTkFont(size=10),
                                               text_color="#9ca3af")
        self.lbl_status_media.pack(side="left", padx=5)

        # ---- CAMPOS MENSAIS (scrollavel) ----
        # Container scrollavel
        canvas_ms = ctk.CTkScrollableFrame(painel_esq, fg_color="transparent", label_text="Pontuação Mensal")
        canvas_ms.pack(fill="both", expand=True, padx=2, pady=3)

        lbl_info = ctk.CTkLabel(canvas_ms,
                                text='Use número ou selecione: ADM FÉRIAS, FÉRIAS, SEM ADM',
                                font=ctk.CTkFont(size=10), text_color="#6b7280")
        lbl_info.pack(padx=10, anchor="w")

        # Cabecalho da tabela mensal
        frm_header = ctk.CTkFrame(canvas_ms, fg_color="#e2e8f0", corner_radius=5)
        frm_header.pack(fill="x", padx=10, pady=(3, 3))
        ctk.CTkLabel(frm_header, text="Mês", width=100, anchor="w",
                      font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=5)
        ctk.CTkLabel(frm_header, text="Pontuação", width=145, anchor="center",
                      font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=5)
        ctk.CTkLabel(frm_header, text="Status", width=140, anchor="center",
                      font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=5)

        # Entradas mensais: cada mes tem Entry para pontuacao + ComboBox para status
        self.entries_pont_mensal = {}
        self.combos_status_mensal = {}

        # Opcoes do combo de status
        opcoes_status = ["(número)"] + VALORES_TEXTO

        for mk, mnome in MESES:
            frm_mes = ctk.CTkFrame(canvas_ms, fg_color="transparent")
            frm_mes.pack(fill="x", padx=10, pady=1)
            ctk.CTkLabel(frm_mes, text=mnome, width=100, anchor="w",
                          font=ctk.CTkFont(size=11)).pack(side="left", padx=5)

            entry_pont = ctk.CTkEntry(frm_mes, width=135, placeholder_text="Pontuação")
            entry_pont.pack(side="left", padx=5)

            combo_status = ctk.CTkComboBox(frm_mes, values=opcoes_status, width=145,
                                            command=lambda val, m=mk: self._ao_selecionar_status(m, val))
            combo_status.set("(número)")
            combo_status.pack(side="left", padx=5)

            self.entries_pont_mensal[mk] = entry_pont
            self.combos_status_mensal[mk] = combo_status

            # Auto-save com debounce
            entry_pont.bind("<KeyRelease>", self._debounce_salvar)

        # Separador
        ctk.CTkLabel(canvas_ms, text="").pack(pady=2)

        # Botoes Salvar / Excluir / Novo / Limpar
        frm_btns = ctk.CTkFrame(canvas_ms, fg_color="transparent")
        frm_btns.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(frm_btns, text="Salvar", fg_color="#22c55e", hover_color="#16a34a",
                       width=120, command=self._salvar_edicao).pack(side="left", padx=5)
        ctk.CTkButton(frm_btns, text="Excluir", fg_color="#ef4444", hover_color="#dc2626",
                       width=120, command=self._excluir_registro).pack(side="left", padx=5)
        ctk.CTkButton(frm_btns, text="Novo", fg_color="#3b82f6", hover_color="#2563eb",
                       width=120, command=self._novo_registro).pack(side="left", padx=5)
        ctk.CTkButton(frm_btns, text="Limpar", fg_color="#f59e0b", hover_color="#d97706",
                       width=120, command=self._limpar_mensal).pack(side="left", padx=5)

        # --- GRAFICO ---
        frm_graf_ctrl = ctk.CTkFrame(painel_dir, fg_color="transparent")
        frm_graf_ctrl.pack(fill="x", padx=10, pady=(10, 0))
        ctk.CTkLabel(frm_graf_ctrl, text="Grafico de Pontuacao",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")

        self.combo_grafico = ctk.CTkSegmentedButton(frm_graf_ctrl,
                                                      values=["Tendência Anual", "Mensal"],
                                                      command=self._ao_mudar_grafico,
                                                      font=ctk.CTkFont(size=12))
        self.combo_grafico.set("Mensal")
        self.combo_grafico.pack(side="right", padx=10)

        self.frame_grafico = ctk.CTkFrame(painel_dir, fg_color="transparent")
        self.frame_grafico.pack(fill="both", expand=True, padx=10, pady=(5, 10))

    def _ao_selecionar_status(self, mes_key, valor):
        """Ao selecionar um status no combo, preenche o entry com o texto e desabilita digitacao."""
        entry = self.entries_pont_mensal[mes_key]
        if valor != "(número)":
            self._atualizando_ui = True
            try:
                entry.configure(state="normal")
                entry.delete(0, "end")
                entry.insert(0, valor)
                entry.configure(state="disabled")
            finally:
                self._atualizando_ui = False
            self._debounce_salvar()
        else:
            self._atualizando_ui = True
            try:
                entry.configure(state="normal")
                val_atual = entry.get().strip()
                if eh_valor_texto(val_atual):
                    entry.delete(0, "end")
            finally:
                self._atualizando_ui = False

    # ---- ABA TABELA ----
    def _criar_aba_tabela(self):
        container = ctk.CTkFrame(self.tab_tabela, corner_radius=10)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview", background="#ffffff", foreground="#333",
                         rowheight=32, font=("Segoe UI", 11))
        style.configure("Custom.Treeview.Heading", font=("Segoe UI", 11, "bold"),
                         background="#e2e8f0", foreground="#333")
        style.map("Custom.Treeview", background=[("selected", "#3b82f6")])

        # Colunas: Estado, Loja, Ano, Jan-Dez, MEDIA, Contrato
        colunas = ["estado", "loja", "ano"] + \
                  [f"pont_{mk}" for mk, _ in MESES] + \
                  ["media", "contrato"]
        self.tree = ttk.Treeview(container, columns=colunas, show="headings",
                                  style="Custom.Treeview", height=22)

        headers = {"estado": ("Estado", 120), "loja": ("Loja", 190), "ano": ("Ano", 55)}
        for mk, mnome in MESES:
            headers[f"pont_{mk}"] = (MESES_CURTOS[MESES.index((mk, mnome))], 88)
        headers["media"] = ("MÉDIA", 90)
        headers["contrato"] = ("Contrato", 170)


        for col in colunas:
            txt, w = headers[col]
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor="center", minwidth=40)

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", self._ao_selecionar_tabela)

    # ---- POPULAR COMBOS ----
    def _popular_combos(self):
        self._atualizando_ui = True
        try:
            estados = obter_estados(self.dados)
            self.combo_estado.configure(values=estados)
            if estados:
                self.combo_estado.set(estados[0])
                lojas = obter_lojas_por_estado(self.dados, estados[0])
                lojas_combo = ["TODAS AS LOJAS"] + lojas
                self.combo_loja.configure(values=lojas_combo)
                self.combo_loja.set("TODAS AS LOJAS")
                # Filtrar anos pelo primeiro estado
                anos = obter_anos_por_estado_loja(self.dados, estados[0])
                anos_str = [str(a) for a in anos]
                self.combo_ano.configure(values=anos_str)
                if anos_str:
                    self.combo_ano.set(anos_str[-1])
                else:
                    self.combo_ano.set("")
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

    def _ao_mudar_estado(self, estado):
        self._atualizando_ui = True
        try:
            lojas = obter_lojas_por_estado(self.dados, estado)
            lojas_combo = ["TODAS AS LOJAS"] + lojas
            self.combo_loja.configure(values=lojas_combo)
            if lojas_combo:
                self.combo_loja.set("TODAS AS LOJAS")
            else:
                self.combo_loja.set("")
            # Filtrar anos pelo estado selecionado
            anos = obter_anos_por_estado_loja(self.dados, estado)
            anos_str = [str(a) for a in anos]
            self.combo_ano.configure(values=anos_str)
            if anos_str:
                self.combo_ano.set(anos_str[-1])
            else:
                self.combo_ano.set("")
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

    # ---- DESENHAR MAPA DO BRASIL (FUNDO DO GRAFICO) ----
    def _desenhar_mapa_fundo(self, ax):
        """Desenha o mapa do Brasil como marca d'agua no fundo do grafico.

        Usa ax.add_patch() individual com transform=ax.transAxes para posicionar
        o mapa em coordenadas de eixo (0-1). PatchCollection nao renderiza com
        transAxes, entao cada poligono e adicionado separadamente.
        O estado selecionado fica em azul semi-transparente; os demais em cinza claro.
        """
        if not HAS_MATPLOTLIB:
            return

        estado_selecionado = self.combo_estado.get().strip()
        sigla_destaque = ESTADO_PARA_SIGLA.get(estado_selecionado, None)

        # Limites do Brasil (lon/lat)
        LON_MIN, LON_MAX = -74, -34
        LAT_MIN, LAT_MAX = -34, 6

        # Posicao do mapa dentro do eixo (coordenadas de eixo 0-1)
        MAP_LEFT = 0.03
        MAP_RIGHT = 0.97
        MAP_BOTTOM = 0.02
        MAP_TOP = 0.98

        # Proporcao do Brasil: lon range=40, lat range=40 ≈ quadrado
        lon_range = LON_MAX - LON_MIN  # 40
        lat_range = LAT_MAX - LAT_MIN  # 40
        aspect = lat_range / lon_range  # ~1.0

        # Calcular area util respeitando aspecto
        axes_w = MAP_RIGHT - MAP_LEFT
        axes_h = MAP_TOP - MAP_BOTTOM
        if aspect >= 1.0:
            map_h = axes_w * aspect
            if map_h > axes_h:
                map_h = axes_h
                map_w = map_h / aspect
            else:
                map_w = axes_w
        else:
            map_w = axes_h / aspect
            if map_w > axes_w:
                map_w = axes_w
                map_h = map_w * aspect
            else:
                map_h = axes_h

        # Centralizar
        offset_x = MAP_LEFT + (axes_w - map_w) / 2
        offset_y = MAP_BOTTOM + (axes_h - map_h) / 2

        def lon_to_ax(lon):
            """Converte longitude para coordenada de eixo (0-1)."""
            return offset_x + (lon - LON_MIN) / (LON_MAX - LON_MIN) * map_w

        def lat_to_ax(lat):
            """Converte latitude para coordenada de eixo (0-1)."""
            return offset_y + (lat - LAT_MIN) / (LAT_MAX - LAT_MIN) * map_h

        # Cores semi-transparentes para marca d'agua
        cor_destaque = (0.231, 0.510, 0.965, 0.12)   # azul rgba alpha=0.12
        cor_outros = (0.88, 0.91, 0.94, 0.10)        # cinza claro alpha=0.10
        cor_borda = (0.70, 0.75, 0.80, 0.20)         # borda alpha=0.20

        # Adicionar cada poligono individualmente com add_patch
        # (PatchCollection nao renderiza com transAxes)
        trans = ax.transAxes
        for sigla, state in MAPA_ESTADOS.items():
            coords_orig = state["coords"]
            coords_conv = [(lon_to_ax(c[0]), lat_to_ax(c[1])) for c in coords_orig]
            is_destaque = (sigla == sigla_destaque)
            fc = cor_destaque if is_destaque else cor_outros
            poly = mpatches.Polygon(coords_conv, closed=True,
                                   facecolor=fc, edgecolor=cor_borda,
                                   linewidth=0.6, zorder=0,
                                   transform=trans)
            ax.add_patch(poly)

        # Rotulo do estado destacado (sigla azul com contorno)
        if sigla_destaque:
            for sigla, state in MAPA_ESTADOS.items():
                if sigla != sigla_destaque:
                    continue
                coords_orig = state["coords"]
                cx_lon = sum(c[0] for c in coords_orig) / len(coords_orig)
                cy_lat = sum(c[1] for c in coords_orig) / len(coords_orig)
                OFFSETS_ROTULO = {"SE": (0.6, -0.2), "PB": (0.4, 0.6), "RN": (0.6, 1.0)}
                ox_lon, oy_lat = OFFSETS_ROTULO.get(sigla, (0, 0))
                lx = lon_to_ax(cx_lon + ox_lon)
                ly = lat_to_ax(cy_lat + oy_lat)
                ax.text(lx, ly, sigla, ha="center", va="center",
                        fontsize=14, fontweight="bold", color="#3b82f6", alpha=0.40,
                        transform=trans,
                        path_effects=[
                            matplotlib.patheffects.withStroke(
                                linewidth=3, foreground="#1e40af")
                        ], zorder=1)

    def _ao_mudar_loja(self, loja):
        # Atualizar anos disponiveis para o estado/loja selecionado
        self._atualizando_ui = True
        try:
            estado = self.combo_estado.get().strip()
            anos = obter_anos_por_estado_loja(self.dados, estado, loja)
            anos_str = [str(a) for a in anos]
            ano_atual = self.combo_ano.get().strip()
            self.combo_ano.configure(values=anos_str)
            # Manter o ano selecionado se ainda existe, senao usar o ultimo disponivel
            if ano_atual in anos_str:
                self.combo_ano.set(ano_atual)
            elif anos_str:
                self.combo_ano.set(anos_str[-1])
            else:
                self.combo_ano.set("")
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

    def _ao_mudar_ano(self, ano):
        if self._atualizando_ui:
            return
        self._carregar_registro()
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

    def _ao_mudar_grafico(self, modo):
        self._atualizar_grafico()

    # ---- CARREGAR REGISTRO NOS CAMPOS ----
    def _carregar_registro(self):
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        self._atualizando_ui = True
        try:
            if loja == "TODAS AS LOJAS":
                self.entry_contrato.delete(0, "end")
                self.entry_contrato.insert(0, "")
                self.lbl_media.configure(text="")
                self.lbl_status_media.configure(text="Selecione uma loja para editar")
                for mk, _ in MESES:
                    self.entries_pont_mensal[mk].configure(state="normal")
                    self.entries_pont_mensal[mk].delete(0, "end")
                    self.combos_status_mensal[mk].set("(número)")
                self._atualizar_ui_anexo("")
                return

            reg = self._buscar_registro(estado, loja, ano_str)
            if reg:
                self.entry_contrato.delete(0, "end")
                self.entry_contrato.insert(0, str(reg.get("contrato", "")))

                media_val = safe_float(reg.get("media", 0))
                if media_val > 0:
                    self.lbl_media.configure(text=formatar_media(media_val))
                    cor = cor_por_pontuacao(media_val)
                    self.lbl_media.configure(text_color=cor)
                else:
                    self.lbl_media.configure(text="")
                    self.lbl_media.configure(text_color="#6b7280")

                # Contar meses com pontuacao vs texto
                num_meses = 0
                txt_meses = 0
                for mk, _ in MESES:
                    val = str(reg.get(f"pont_{mk}", "")).strip()
                    if val:
                        if eh_valor_texto(val):
                            txt_meses += 1
                        else:
                            num_meses += 1
                status_txt = f"{num_meses} meses com pontuação, {txt_meses} com status textual"
                self.lbl_status_media.configure(text=status_txt)

                # Preencher campos mensais
                for mk, _ in MESES:
                    val = str(reg.get(f"pont_{mk}", "")).strip()
                    entry = self.entries_pont_mensal[mk]
                    combo = self.combos_status_mensal[mk]

                    entry.configure(state="normal")
                    entry.delete(0, "end")

                    if not val:
                        combo.set("(número)")
                    elif eh_valor_texto(val):
                        entry.insert(0, val)
                        entry.configure(state="disabled")
                        combo.set(val)
                    else:
                        # Garantir 2 casas decimais na exibicao
                        try:
                            f = float(val.replace(",", "."))
                            entry.insert(0, formatar_pontuacao(f))
                        except (ValueError, TypeError):
                            entry.insert(0, val)
                        combo.set("(número)")

                # Anexo
                self._atualizar_ui_anexo(reg.get("anexo", ""))
            else:
                self.entry_contrato.delete(0, "end")
                self.entry_contrato.insert(0, "ASSAÍ - ATACADISTA")
                self.lbl_media.configure(text="")
                self.lbl_media.configure(text_color="#6b7280")
                self.lbl_status_media.configure(text="Registro novo · sem dados")
                for mk, _ in MESES:
                    self.entries_pont_mensal[mk].configure(state="normal")
                    self.entries_pont_mensal[mk].delete(0, "end")
                    self.combos_status_mensal[mk].set("(número)")
                self._atualizar_ui_anexo("")
        finally:
            self._atualizando_ui = False

    def _buscar_registro(self, estado, loja, ano_str):
        try:
            ano = int(ano_str)
        except (ValueError, TypeError):
            return None
        for r in self.dados:
            if (str(r.get("estado", "")).strip().upper() == estado.upper() and
                str(r.get("loja", "")).strip().upper() == loja.upper() and
                r.get("ano") == ano):
                return r
        return None

    # ---- ATUALIZAR MEDIA DA UI ----
    def _recalcular_media_ui(self):
        """Recalcula a MEDIA a partir dos campos mensais na tela."""
        if self._atualizando_ui:
            return
        valores = []
        txt_count = 0
        for mk, _ in MESES:
            entry = self.entries_pont_mensal[mk]
            try:
                val = entry.get().strip()
            except Exception:
                val = ""
            if val and eh_valor_texto(val):
                txt_count += 1
            elif val:
                try:
                    f = float(val.replace(",", "."))
                    if 0 <= f <= 100:
                        valores.append(f)
                except (ValueError, TypeError):
                    pass

        if valores:
            media = round(sum(valores) / len(valores), 2)
            self.lbl_media.configure(text=formatar_media(media))
            self.lbl_media.configure(text_color=cor_por_pontuacao(media))
        else:
            self.lbl_media.configure(text="")
            self.lbl_media.configure(text_color="#6b7280")

        num_count = len(valores)
        self.lbl_status_media.configure(
            text=f"{num_count} meses com pontuação, {txt_count} com status textual"
        )

    # ---- ATUALIZAR TABELA ----
    def _atualizar_tabela(self):
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        for item in self.tree.get_children():
            self.tree.delete(item)

        todas_lojas = (loja == "TODAS AS LOJAS")

        # Filtrar por ano se selecionado
        filtrar_ano = False
        ano_filtro = None
        if ano_str:
            try:
                ano_filtro = int(ano_str)
                filtrar_ano = True
            except (ValueError, TypeError):
                filtrar_ano = False

        for r in self.dados:
            if estado and str(r.get("estado", "")).strip().upper() != estado.upper():
                continue
            if not todas_lojas and loja and str(r.get("loja", "")).strip().upper() != loja.upper():
                continue
            if filtrar_ano and r.get("ano") != ano_filtro:
                continue

            vals = [
                r.get("estado", ""),
                r.get("loja", ""),
                r.get("ano", ""),
            ]
            # Adicionar pontuacao mensal — sempre com 2 casas decimais
            for mk, _ in MESES:
                val = str(r.get(f"pont_{mk}", "")).strip()
                if val:
                    if eh_valor_texto(val):
                        vals.append(val)
                    else:
                        try:
                            f = float(val.replace(",", "."))
                            vals.append(formatar_pontuacao(f))
                        except (ValueError, TypeError):
                            vals.append(val)
                else:
                    vals.append("")

            # MEDIA — 2 casas decimais
            media_val = safe_float(r.get("media", 0))
            vals.append(formatar_media(media_val) if media_val > 0 else "")

            # Contrato
            vals.append(r.get("contrato", ""))

            tag = tag_por_pontuacao(media_val)
            self.tree.insert("", "end", values=vals, tags=(tag,))

        self.tree.tag_configure("excelente", background="#dcfce7")  # verde >= 90
        self.tree.tag_configure("bom", background="#fef9c3")  # amarelo >= 70
        self.tree.tag_configure("baixo", background="#fee2e2")  # vermelho < 70
        self.tree.tag_configure("vazio", background="#ffffff")
        self.tree.tag_configure("texto", background="#f3f4f6")  # cinza para texto

        todas_lojas = (loja == "TODAS AS LOJAS")
        filtrados = [r for r in self.dados
                     if str(r.get("estado", "")).strip().upper() == estado.upper()
                     and (todas_lojas or str(r.get("loja", "")).strip().upper() == loja.upper())
                     and (not filtrar_ano or r.get("ano") == ano_filtro)]
        n = len(filtrados)
        ano_txt = f" │ {ano_str}" if ano_str else ""
        if todas_lojas:
            self.lbl_contador.configure(text=f"{n} registros │ Todas as Lojas ({estado}){ano_txt}")
        else:
            self.lbl_contador.configure(text=f"{n} registros │ {loja} ({estado}){ano_txt}")

    # ---- ATUALIZAR GRAFICO ----
    def _atualizar_grafico(self):
        if not HAS_MATPLOTLIB:
            return

        modo = self.combo_grafico.get() if hasattr(self, 'combo_grafico') else "Mensal"
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        todas_lojas = (loja == "TODAS AS LOJAS")

        for w in self.frame_grafico.winfo_children():
            w.destroy()

        if modo == "Mensal":
            self._gerar_grafico_mensal(estado, loja, todas_lojas)
        else:
            self._gerar_grafico_anual(estado, loja, todas_lojas)

    def _gerar_grafico_mensal(self, estado, loja, todas_lojas):
        """Grafico de barras mostrando pontuacao mensal para o ano selecionado."""
        try:
            ano = int(self.combo_ano.get().strip())
        except (ValueError, TypeError):
            ctk.CTkLabel(self.frame_grafico, text="Ano inválido.",
                         font=ctk.CTkFont(size=14), text_color="gray").pack(pady=30)
            return

        if todas_lojas:
            registros = [r for r in self.dados
                         if str(r.get("estado", "")).strip().upper() == estado.upper()
                         and r.get("ano") == ano]
        else:
            registros = [r for r in self.dados
                         if str(r.get("estado", "")).strip().upper() == estado.upper()
                         and str(r.get("loja", "")).strip().upper() == loja.upper()
                         and r.get("ano") == ano]

        if not registros:
            ctk.CTkLabel(self.frame_grafico, text=f"Nenhum dado encontrado para {ano}.",
                         font=ctk.CTkFont(size=14), text_color="gray").pack(pady=30)
            return

        # Agregar pontuacoes mensais (media dos registros se TODAS AS LOJAS)
        ponts_agregado = {}
        contagens = {}
        textos_agregado = {}
        for r in registros:
            for mk, _ in MESES:
                val = str(r.get(f"pont_{mk}", "")).strip()
                if val and eh_valor_texto(val):
                    if mk not in textos_agregado:
                        textos_agregado[mk] = []
                    textos_agregado[mk].append(val)
                elif val:
                    try:
                        f = float(val.replace(",", "."))
                        ponts_agregado[mk] = ponts_agregado.get(mk, 0) + f
                        contagens[mk] = contagens.get(mk, 0) + 1
                    except (ValueError, TypeError):
                        pass

        # Calcular media agregada por mes
        ponts_mensal = []
        for mk, _ in MESES:
            if mk in ponts_agregado and mk in contagens:
                ponts_mensal.append(ponts_agregado[mk] / contagens[mk])
            elif mk in ponts_agregado:
                ponts_mensal.append(ponts_agregado[mk])
            else:
                ponts_mensal.append(0)

        tem_valores = any(p > 0 for p in ponts_mensal)
        if not tem_valores:
            ctk.CTkLabel(self.frame_grafico,
                         text="Sem dados numericos para exibir.\nInsira pontuacoes nos campos mensais.",
                         font=ctk.CTkFont(size=14), text_color="gray").pack(pady=50)
            return

        if todas_lojas:
            titulo = f"Média das Lojas — {estado} ({ano})"
        else:
            titulo = f"{loja} — {ano}"

        fig = Figure(figsize=(8, 4.5), dpi=100, facecolor="#f8fafc")
        ax = fig.add_subplot(111)

        x_pos = np.arange(12) if HAS_NUMPY else list(range(12))

        # Spline suave para a linha
        usar_spline = False
        pontos_validos = [(i, p) for i, p in enumerate(ponts_mensal) if p > 0]
        if HAS_SCIPY and HAS_NUMPY and len(pontos_validos) >= 3:
            try:
                xi = np.array([pv[0] for pv in pontos_validos])
                yi = np.array([pv[1] for pv in pontos_validos])
                grau = min(3, len(pontos_validos) - 1)
                spl = make_interp_spline(xi, yi, k=grau)
                x_smooth = np.linspace(xi.min(), xi.max(), 200)
                _ = spl(x_smooth[:3])
                usar_spline = True
            except Exception:
                usar_spline = False

        # Cor da linha baseada na media
        media_val = sum(ponts_mensal) / max(1, sum(1 for p in ponts_mensal if p > 0))
        if media_val >= 90:
            cor_linha = "#22c55e"
        elif media_val >= 70:
            cor_linha = "#3b82f6"
        else:
            cor_linha = "#ef4444"

        if usar_spline:
            ax.plot(x_smooth, spl(x_smooth), color=cor_linha, linewidth=2.5)
            ax.scatter(x_pos, ponts_mensal, color=cor_linha, s=70, zorder=5,
                       edgecolors="white", linewidths=1.5)
        else:
            ax.plot(x_pos, ponts_mensal, color=cor_linha, linewidth=2.5, marker="o",
                    markersize=8, markeredgecolor="white", markeredgewidth=2)

        # Preencher area sob a curva
        if HAS_NUMPY:
            x_arr = np.array(x_pos)
            y_arr = np.array(ponts_mensal)
            ax.fill_between(x_arr, y_arr, alpha=0.08, color=cor_linha)

        # Anotacoes dos valores (tamanho 12)
        for i, p in enumerate(ponts_mensal):
            if p > 0:
                ax.annotate(formatar_pontuacao(p), (i, p), textcoords="offset points",
                            xytext=(0, 12), ha="center", fontsize=12, fontweight="bold",
                            color=cor_linha)
            # Mostrar texto de status se houver
            mk = MESES[i][0]
            if mk in textos_agregado and textos_agregado[mk]:
                txt = textos_agregado[mk][0]
                ax.annotate(txt, (i, 2), ha="center", fontsize=9, color="#6b7280",
                            fontstyle="italic")

        ax.set_title(titulo, fontsize=14, fontweight="bold", pad=10)
        ax.set_xlabel("Mês", fontsize=12)
        ax.set_xticks(x_pos)
        ax.set_xticklabels(MESES_CURTOS, fontsize=10)
        ax.set_ylim(0, 110)
        ax.set_facecolor("#ffffff")
        # Mapa do Brasil como marca d'agua no fundo (usa transAxes)
        self._desenhar_mapa_fundo(ax)
        # Reaplicar limites apos mapa
        ax.set_xlim(-0.5, 11.5)
        ax.set_ylim(0, 110)
        # Remover eixo Y completamente
        ax.set_yticks([])
        ax.spines["left"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["top"].set_visible(False)
        # Sem linha de grade
        ax.grid(False)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def _gerar_grafico_anual(self, estado, loja, todas_lojas):
        """Grafico de tendencia: MEDIA por ano."""
        if todas_lojas:
            registros = [r for r in self.dados
                         if str(r.get("estado", "")).strip().upper() == estado.upper()]
        else:
            registros = [r for r in self.dados
                         if str(r.get("estado", "")).strip().upper() == estado.upper()
                         and str(r.get("loja", "")).strip().upper() == loja.upper()]

        if not registros:
            ctk.CTkLabel(self.frame_grafico, text="Nenhum dado encontrado.",
                         font=ctk.CTkFont(size=14), text_color="gray").pack(pady=30)
            return

        # Calcular MEDIA agregada por ano
        soma_media = defaultdict(float)
        contagem = defaultdict(int)
        for r in registros:
            a = int(r.get("ano", 0))
            m = safe_float(r.get("media", 0))
            if m > 0:
                soma_media[a] += m
                contagem[a] += 1

        anos = sorted(soma_media.keys())
        medias = [soma_media[a] / contagem[a] if contagem[a] > 0 else 0 for a in anos]

        tem_valores = any(m > 0 for m in medias)
        if not tem_valores:
            ctk.CTkLabel(self.frame_grafico,
                         text="Sem dados numericos para exibir.\nInsira pontuacoes nos campos mensais.",
                         font=ctk.CTkFont(size=14), text_color="gray").pack(pady=50)
            return

        if todas_lojas:
            titulo = f"Tendência Anual — {estado} (Todas as Lojas)"
        else:
            titulo = f"Tendência Anual — {loja}"

        fig = Figure(figsize=(8, 4.5), dpi=100, facecolor="#f8fafc")
        ax = fig.add_subplot(111)

        if HAS_NUMPY:
            x = np.array(anos)
            y = np.array(medias)
        else:
            x = anos
            y = medias

        # Spline suave
        usar_spline = False
        if HAS_SCIPY and HAS_NUMPY and len(anos) >= 3:
            anos_unicos = len(set(anos))
            if anos_unicos == len(anos):
                try:
                    x_smooth = np.linspace(x.min(), x.max(), 200)
                    grau = min(3, len(anos) - 1)
                    spl = make_interp_spline(x, y, k=grau)
                    _ = spl(x_smooth[:3])
                    usar_spline = True
                except Exception:
                    usar_spline = False

        # Cor da linha baseada na media
        media_anual = sum(medias) / max(1, len([m for m in medias if m > 0]))
        if media_anual >= 90:
            cor_linha = "#22c55e"
        elif media_anual >= 70:
            cor_linha = "#3b82f6"
        else:
            cor_linha = "#ef4444"

        if usar_spline:
            ax.plot(x_smooth, spl(x_smooth), color=cor_linha, linewidth=2.5)
            ax.scatter(x, y, color=cor_linha, s=70, zorder=5, edgecolors="white", linewidths=1.5)
        else:
            ax.plot(x, y, color=cor_linha, linewidth=2.5, marker="o", markersize=8,
                    markeredgecolor="white", markeredgewidth=2)

        # Anotacoes (tamanho 12)
        for a_val, m_val in zip(anos, y):
            if m_val > 0:
                ax.annotate(formatar_media(m_val), (a_val, m_val), textcoords="offset points",
                            xytext=(0, 14), ha="center", fontsize=12, color=cor_linha,
                            fontweight="bold")

        # Preencher area sob a curva
        if HAS_NUMPY and len(anos) >= 2:
            ax.fill_between(x, y, alpha=0.08, color=cor_linha)

        ax.set_title(titulo, fontsize=14, fontweight="bold", pad=10)
        ax.set_xlabel("Ano", fontsize=12)
        ax.set_ylim(0, 110)
        ax.set_facecolor("#ffffff")
        # Mapa do Brasil como marca d'agua no fundo (usa transAxes)
        self._desenhar_mapa_fundo(ax)
        # Reaplicar limites apos mapa
        ax.set_xlim(min(anos) - 0.5, max(anos) + 0.5)
        ax.set_ylim(0, 110)
        # Remover eixo Y completamente
        ax.set_yticks([])
        ax.spines["left"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["top"].set_visible(False)
        # Sem linha de grade
        ax.grid(False)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    # ---- ANEXO ----
    def _atualizar_ui_anexo(self, caminho):
        """Atualiza os botoes e label do anexo conforme o caminho."""
        tem_anexo = bool(caminho and os.path.exists(caminho))
        if tem_anexo:
            nome_arq = os.path.basename(caminho)
            self.lbl_anexo_status.configure(text=f"{nome_arq}", text_color="#4338ca")
            self.btn_visualizar.configure(state="normal")
            self.btn_baixar.configure(state="normal")
            self.btn_remover_anexo.configure(state="normal")
        else:
            self.lbl_anexo_status.configure(text="Nenhum anexo", text_color="#6b7280")
            self.btn_visualizar.configure(state="disabled")
            self.btn_baixar.configure(state="disabled")
            self.btn_remover_anexo.configure(state="disabled")

    def _anexar_arquivo(self):
        """Permite selecionar um arquivo para anexar ao registro atual."""
        loja = self.combo_loja.get().strip()
        if loja == "TODAS AS LOJAS":
            messagebox.showinfo("Aviso", "Selecione uma loja específica para anexar.")
            return

        estado = self.combo_estado.get().strip()
        ano_str = self.combo_ano.get().strip()

        caminho_orig = filedialog.askopenfilename(
            title="Selecionar Anexo da Pontuação",
            filetypes=[
                ("Imagens", "*.png *.jpg *.jpeg *.bmp *.gif *.tiff *.webp"),
                ("PDF", "*.pdf"),
                ("Planilhas", "*.xlsx *.xls *.csv"),
                ("Documentos", "*.docx *.doc *.txt"),
                ("Todos os arquivos", "*.*")
            ]
        )
        if not caminho_orig:
            return

        # Copiar para pasta de anexos com nome seguro
        nome_base = _nome_arquivo_anexo(estado, loja, ano_str)
        ext = os.path.splitext(caminho_orig)[1]
        nome_destino = f"{nome_base}{ext}"

        # Se ja existe arquivo com mesmo nome, adicionar hash curto
        caminho_dest = os.path.join(PASTA_ANEXOS, nome_destino)
        if os.path.exists(caminho_dest) and os.path.abspath(caminho_orig) != os.path.abspath(caminho_dest):
            # Adicionar timestamp para evitar conflito
            ts = datetime.now().strftime("%H%M%S")
            nome_destino = f"{nome_base}_{ts}{ext}"
            caminho_dest = os.path.join(PASTA_ANEXOS, nome_destino)

        try:
            shutil.copy2(caminho_orig, caminho_dest)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao copiar anexo:\n{e}")
            return

        # Atualizar registro
        reg = self._buscar_registro(estado, loja, ano_str)
        if reg:
            # Remover anexo antigo se diferente
            anexo_antigo = reg.get("anexo", "")
            if anexo_antigo and anexo_antigo != caminho_dest and os.path.exists(anexo_antigo):
                try:
                    os.remove(anexo_antigo)
                except Exception:
                    pass
            reg["anexo"] = caminho_dest
        else:
            # Criar registro novo se nao existe
            try:
                ano = int(ano_str)
            except (ValueError, TypeError):
                ano = 2026
            novo = _reg(estado, loja, "ASSAÍ - ATACADISTA", ano)
            novo["anexo"] = caminho_dest
            calcular_media(novo)
            self.dados.append(novo)

        salvar_dados_json(self.dados)
        self._atualizar_ui_anexo(caminho_dest)
        self._atualizar_tabela()

    def _visualizar_anexo(self):
        """Abre o anexo com o aplicativo padrao do sistema."""
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        reg = self._buscar_registro(estado, loja, ano_str)
        if not reg:
            messagebox.showinfo("Aviso", "Nenhum registro selecionado.")
            return

        caminho = reg.get("anexo", "")
        if not caminho or not os.path.exists(caminho):
            messagebox.showinfo("Aviso", "Anexo não encontrado.\nO arquivo pode ter sido movido ou excluído.")
            self._atualizar_ui_anexo("")
            return

        # Se for imagem, abrir janela de visualizacao interna
        ext_lower = os.path.splitext(caminho)[1].lower()
        if ext_lower in [".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff", ".webp"]:
            self._visualizar_imagem(caminho)
        else:
            # Para outros tipos (PDF, xlsx, etc), abrir com app padrao
            _abrir_arquivo(caminho)

    def _visualizar_imagem(self, caminho):
        """Abre janela para visualizar imagem do anexo."""
        if self._janela_anexo and self._janela_anexo.winfo_exists():
            self._janela_anexo.destroy()

        self._janela_anexo = ctk.CTkToplevel(self)
        self._janela_anexo.title("Visualizar Anexo")
        self._janela_anexo.geometry("900x700")
        self._janela_anexo.minsize(400, 300)

        nome_arq = os.path.basename(caminho)

        # Topo com info
        frm_top = ctk.CTkFrame(self._janela_anexo, fg_color="transparent")
        frm_top.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(frm_top, text=f"{nome_arq}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(side="left")
        ctk.CTkButton(frm_top, text="Baixar como...", fg_color="#6366f1",
                      hover_color="#4f46e5", width=130,
                      command=lambda: self._baixar_anexo_como(caminho)).pack(side="right", padx=5)
        ctk.CTkButton(frm_top, text="Fechar", fg_color="#6b7280",
                      hover_color="#4b5563", width=80,
                      command=self._janela_anexo.destroy).pack(side="right", padx=5)

        # Container de imagem
        frm_img = ctk.CTkFrame(self._janela_anexo, fg_color="#f8fafc", corner_radius=8)
        frm_img.pack(fill="both", expand=True, padx=10, pady=5)

        try:
            if HAS_PIL:
                img = Image.open(caminho)
                # Redimensionar mantendo proporcao
                max_w = 850
                max_h = 600
                w, h = img.size
                ratio = min(max_w / w, max_h / h, 1.0)
                new_w = int(w * ratio)
                new_h = int(h * ratio)
                img = img.resize((new_w, new_h), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                lbl_img = tk.Label(frm_img, image=photo, bg="#f8fafc")
                lbl_img.image = photo  # manter referencia
                lbl_img.pack(padx=10, pady=10)
            else:
                # Sem PIL: abrir com app padrao
                _abrir_arquivo(caminho)
                ctk.CTkLabel(frm_img, text=f"Arquivo: {nome_arq}\n(Abra com o aplicativo padrão)",
                             font=ctk.CTkFont(size=13), text_color="#6b7280").pack(pady=50)
        except Exception as e:
            ctk.CTkLabel(frm_img, text=f"Erro ao carregar imagem:\n{e}",
                         font=ctk.CTkFont(size=13), text_color="#ef4444").pack(pady=50)

    def _baixar_anexo(self):
        """Salva o anexo em local escolhido pelo usuario."""
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        reg = self._buscar_registro(estado, loja, ano_str)
        if not reg:
            messagebox.showinfo("Aviso", "Nenhum registro selecionado.")
            return

        caminho = reg.get("anexo", "")
        if not caminho or not os.path.exists(caminho):
            messagebox.showinfo("Aviso", "Anexo não encontrado.")
            self._atualizar_ui_anexo("")
            return

        self._baixar_anexo_como(caminho)

    def _baixar_anexo_como(self, caminho_origem):
        """Copia o anexo para um local escolhido pelo usuario."""
        nome_arq = os.path.basename(caminho_origem)
        caminho_dest = filedialog.asksaveasfilename(
            defaultextension=os.path.splitext(nome_arq)[1],
            filetypes=[("Todos os arquivos", "*.*")],
            initialfile=nome_arq,
            title="Salvar Anexo Como..."
        )
        if not caminho_dest:
            return
        try:
            shutil.copy2(caminho_origem, caminho_dest)
            messagebox.showinfo("Sucesso", f"Anexo salvo em:\n{caminho_dest}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar anexo:\n{e}")

    def _remover_anexo(self):
        """Remove o anexo do registro atual."""
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        reg = self._buscar_registro(estado, loja, ano_str)
        if not reg:
            return

        anexo_path = reg.get("anexo", "")
        if not anexo_path:
            return

        if messagebox.askyesno("Confirmar", "Deseja remover o anexo deste registro?"):
            # Deletar arquivo fisico
            if os.path.exists(anexo_path):
                try:
                    os.remove(anexo_path)
                except Exception:
                    pass
            reg["anexo"] = ""
            salvar_dados_json(self.dados)
            self._atualizar_ui_anexo("")
            self._atualizar_tabela()

    # ---- DEBOUNCE AUTO-SAVE ----
    def _debounce_salvar(self, event=None):
        if self._atualizando_ui:
            return
        if self._debounce_id:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(900, self._salvar_edicao)

    # ---- SALVAR EDICAO ----
    def _salvar_edicao(self):
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        if not estado or not loja or loja == "TODAS AS LOJAS":
            return

        try:
            ano = int(ano_str)
        except (ValueError, TypeError):
            return

        contrato = self.entry_contrato.get().strip() or "ASSAÍ - ATACADISTA"

        # Ler campos mensais de pontuacao
        pont_mensal = {}
        for mk, _ in MESES:
            entry = self.entries_pont_mensal[mk]
            try:
                val = entry.get().strip()
            except Exception:
                val = ""
            # Validar: numero entre 0-100 ou texto de status
            if val and eh_valor_texto(val):
                pont_mensal[mk] = val
            elif val:
                try:
                    f = float(val.replace(",", "."))
                    if 0 <= f <= 100:
                        # Sempre armazenar com 2 casas decimais
                        pont_mensal[mk] = formatar_pontuacao(f)
                    else:
                        pont_mensal[mk] = ""
                except (ValueError, TypeError):
                    pont_mensal[mk] = val
            else:
                pont_mensal[mk] = ""

        reg = self._buscar_registro(estado, loja, ano_str)
        if reg:
            reg["contrato"] = contrato
            for mk, _ in MESES:
                reg[f"pont_{mk}"] = pont_mensal[mk]
            calcular_media(reg)
        else:
            novo = {"estado": estado, "loja": loja, "contrato": contrato, "ano": ano}
            for mk, _ in MESES:
                novo[f"pont_{mk}"] = pont_mensal[mk]
            novo["anexo"] = ""
            calcular_media(novo)
            self.dados.append(novo)

        # Atualizar MEDIA na UI
        reg2 = self._buscar_registro(estado, loja, ano_str)
        if reg2:
            self._atualizando_ui = True
            try:
                media_val = safe_float(reg2.get("media", 0))
                if media_val > 0:
                    self.lbl_media.configure(text=formatar_media(media_val))
                    self.lbl_media.configure(text_color=cor_por_pontuacao(media_val))
                else:
                    self.lbl_media.configure(text="")
                    self.lbl_media.configure(text_color="#6b7280")

                # Atualizar entries com valor formatado (precisao completa)
                # Garante que o campo mostra '84,92' ao inves de '84.9'
                for mk, _ in MESES:
                    entry = self.entries_pont_mensal[mk]
                    combo = self.combos_status_mensal[mk]
                    saved = reg2.get(f"pont_{mk}", "")
                    if saved and not eh_valor_texto(saved):
                        try:
                            f_val = float(saved.replace(",", "."))
                            fmt = formatar_pontuacao(f_val)
                            if entry.cget("state") != "disabled":
                                cur = entry.get().strip()
                                if cur != fmt:
                                    pos = entry.index("insert")
                                    entry.delete(0, "end")
                                    entry.insert(0, fmt)
                                    # Manter cursor na posicao correta
                                    try:
                                        entry.icursor(pos)
                                    except Exception:
                                        pass
                        except (ValueError, TypeError):
                            pass
            finally:
                self._atualizando_ui = False

        salvar_dados_json(self.dados)
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

    # ---- LIMPAR MENSAL ----
    def _limpar_mensal(self):
        loja = self.combo_loja.get().strip()
        if loja == "TODAS AS LOJAS":
            messagebox.showinfo("Aviso", "Selecione uma loja específica.")
            return
        self._atualizando_ui = True
        try:
            for mk, _ in MESES:
                self.entries_pont_mensal[mk].configure(state="normal")
                self.entries_pont_mensal[mk].delete(0, "end")
                self.combos_status_mensal[mk].set("(número)")
        finally:
            self._atualizando_ui = False
        self._salvar_edicao()

    # ---- REGISTRAR NOVO ESTADO ----
    def _novo_estado(self):
        dlg = ctk.CTkInputDialog(text="Nome do novo Estado:", title="Registrar Novo Estado")
        nome = dlg.get_input()
        if nome is None or not nome.strip():
            return
        nome = nome.strip().upper()
        estados_atuais = self.combo_estado.cget("values")
        if nome in estados_atuais:
            messagebox.showinfo("Aviso", f"Estado \"{nome}\" já existe.")
            self.combo_estado.set(nome)
            self._ao_mudar_estado(nome)
            return
        # Adicionar ao combo e selecionar
        self._atualizando_ui = True
        try:
            novos = list(estados_atuais) + [nome]
            self.combo_estado.configure(values=novos)
            self.combo_estado.set(nome)
        finally:
            self._atualizando_ui = False
        # Como é estado novo, só tem TODAS AS LOJAS; pedir nome da primeira loja
        dlg2 = ctk.CTkInputDialog(text=f"Nome da primeira loja em {nome}:", title="Registrar Primeira Loja")
        loja_nome = dlg2.get_input()
        if loja_nome is None or not loja_nome.strip():
            loja_nome = f"ASSAÍ - {nome}"
        loja_nome = loja_nome.strip().upper()
        # Atualizar combo_loja para mostrar a nova loja
        self._atualizando_ui = True
        try:
            self.combo_loja.configure(values=["TODAS AS LOJAS", loja_nome])
            self.combo_loja.set(loja_nome)
        finally:
            self._atualizando_ui = False
        # Preencher anos disponíveis e selecionar ano atual
        ano_atual = str(datetime.now().year)
        self._atualizando_ui = True
        try:
            self.combo_ano.configure(values=[ano_atual])
            self.combo_ano.set(ano_atual)
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_grafico()
        self._atualizar_resumo()

    # ---- REGISTRAR NOVA LOJA ----
    def _nova_loja(self):
        estado = self.combo_estado.get().strip()
        if not estado:
            messagebox.showinfo("Aviso", "Selecione um Estado primeiro.")
            return
        dlg = ctk.CTkInputDialog(text=f"Nome da nova loja em {estado}:", title="Registrar Nova Loja")
        nome = dlg.get_input()
        if nome is None or not nome.strip():
            return
        nome = nome.strip().upper()
        lojas_atuais = list(self.combo_loja.cget("values"))
        if nome in lojas_atuais:
            messagebox.showinfo("Aviso", f"Loja \"{nome}\" já existe.")
            self.combo_loja.set(nome)
            self._ao_mudar_loja(nome)
            return
        # Inserir antes de "TODAS AS LOJAS" se existir
        novas = []
        inserido = False
        for v in lojas_atuais:
            if v == "TODAS AS LOJAS" and not inserido:
                novas.append(nome)
                inserido = True
            novas.append(v)
        if not inserido:
            novas.append(nome)
        self._atualizando_ui = True
        try:
            self.combo_loja.configure(values=novas)
            self.combo_loja.set(nome)
        finally:
            self._atualizando_ui = False
        # Atualizar anos para esta nova combinação
        anos = obter_anos_por_estado_loja(self.dados, estado, nome)
        ano_atual_int = datetime.now().year
        if ano_atual_int not in anos:
            anos.append(ano_atual_int)
            anos.sort(reverse=True)
        anos_str = [str(a) for a in anos]
        self._atualizando_ui = True
        try:
            self.combo_ano.configure(values=anos_str)
            self.combo_ano.set(str(ano_atual_int))
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_grafico()
        self._atualizar_resumo()

    # ---- REGISTRAR NOVO ANO ----
    def _novo_ano(self):
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        if not estado or not loja or loja == "TODAS AS LOJAS":
            messagebox.showinfo("Aviso", "Selecione um Estado e uma loja específica primeiro.")
            return
        dlg = ctk.CTkInputDialog(text="Novo ano (ex: 2027):", title="Registrar Novo Ano")
        ano_str = dlg.get_input()
        if ano_str is None or not ano_str.strip():
            return
        ano_str = ano_str.strip()
        try:
            ano_int = int(ano_str)
            if ano_int < 2000 or ano_int > 2100:
                messagebox.showwarning("Aviso", "Ano inválido. Use um valor entre 2000 e 2100.")
                return
        except ValueError:
            messagebox.showwarning("Aviso", "Digite um ano válido (ex: 2027).")
            return
        anos_atuais = list(self.combo_ano.cget("values"))
        if ano_str not in anos_atuais:
            anos_atuais.append(ano_str)
            anos_atuais.sort(reverse=True)
        self._atualizando_ui = True
        try:
            self.combo_ano.configure(values=anos_atuais)
            self.combo_ano.set(ano_str)
        finally:
            self._atualizando_ui = False
        self._carregar_registro()
        self._atualizar_grafico()
        self._atualizar_resumo()

    # ---- NOVO REGISTRO ----
    def _novo_registro(self):
        estado = self.combo_estado.get().strip() or "MANAUS"
        loja = self.combo_loja.get().strip()
        if not loja or loja == "TODAS AS LOJAS":
            messagebox.showinfo("Aviso", "Selecione uma loja específica antes de criar um novo registro.")
            return
        ano_str = self.combo_ano.get().strip() or "2026"
        try:
            ano = int(ano_str)
        except ValueError:
            ano = 2026

        if self._buscar_registro(estado, loja, ano_str):
            messagebox.showinfo("Aviso", "Este registro já existe. Edite os campos e clique Salvar.")
            return

        novo = _reg(estado, loja, "ASSAÍ - ATACADISTA", ano)
        calcular_media(novo)
        self.dados.append(novo)
        salvar_dados_json(self.dados)
        self._carregar_registro()
        self._atualizar_tabela()
        self._atualizar_grafico()

    # ---- EXCLUIR REGISTRO ----
    def _excluir_registro(self):
        estado = self.combo_estado.get().strip()
        loja = self.combo_loja.get().strip()
        ano_str = self.combo_ano.get().strip()

        if not estado or not loja or loja == "TODAS AS LOJAS":
            return

        reg = self._buscar_registro(estado, loja, ano_str)
        if not reg:
            return

        if messagebox.askyesno("Confirmar", f"Excluir registro {loja} - {ano_str}?"):
            # Remover anexo se existir
            anexo = reg.get("anexo", "")
            if anexo and os.path.exists(anexo):
                try:
                    os.remove(anexo)
                except Exception:
                    pass
            self.dados.remove(reg)
            salvar_dados_json(self.dados)
            self._carregar_registro()
            self._atualizar_tabela()
            self._atualizar_grafico()
            self._atualizar_resumo()

    # ---- SELECIONAR NA TABELA ----
    def _ao_selecionar_tabela(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        vals = item["values"]
        if len(vals) >= 3:
            self.combo_estado.set(vals[0])
            lojas = obter_lojas_por_estado(self.dados, str(vals[0]))
            lojas_combo = ["TODAS AS LOJAS"] + lojas
            self.combo_loja.configure(values=lojas_combo)
            self.combo_loja.set(vals[1])
            self.combo_ano.set(str(vals[2]))
            self._carregar_registro()
            self._atualizar_grafico()
            self._atualizar_resumo()

    # ---- EXPORTAR EXCEL ----
    def _exportar_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Erro", "openpyxl não está instalado.\nRode: pip install openpyxl")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"checklist_assai_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not caminho:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Checklist Pontuação"

            # Cabecalho
            cabecalho = COLUNAS_EXPORTACAO

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2d2d2d", end_color="2d2d2d",
                                       fill_type="solid")
            for col_idx, h in enumerate(cabecalho, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            excelente_fill = PatternFill(start_color="dcfce7", end_color="dcfce7",
                                          fill_type="solid")
            bom_fill = PatternFill(start_color="fef9c3", end_color="fef9c3",
                                    fill_type="solid")
            baixo_fill = PatternFill(start_color="fee2e2", end_color="fee2e2",
                                      fill_type="solid")
            texto_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6",
                                      fill_type="solid")

            for row_idx, r in enumerate(self.dados, 2):
                col = 1
                # ANO
                ws.cell(row=row_idx, column=col, value=r.get("ano", "")); col += 1
                # ESTADOS
                ws.cell(row=row_idx, column=col, value=r.get("estado", "")); col += 1
                # LOJAS
                ws.cell(row=row_idx, column=col, value=r.get("loja", "")); col += 1

                # Jan-Dez (pontuacao mensal)
                for mk, _ in MESES:
                    val = str(r.get(f"pont_{mk}", "")).strip()
                    if val and not eh_valor_texto(val):
                        try:
                            ws.cell(row=row_idx, column=col,
                                    value=float(val.replace(",", ".")))
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=col, value=val)
                    elif val:
                        ws.cell(row=row_idx, column=col, value=val)
                    else:
                        ws.cell(row=row_idx, column=col, value="")
                    col += 1

                # MEDIA
                media_val = safe_float(r.get("media", 0))
                ws.cell(row=row_idx, column=col, value=media_val if media_val > 0 else "")
                col += 1

                # CONTRATO
                ws.cell(row=row_idx, column=col, value=r.get("contrato", ""))

                # Colorir MEDIA
                tag = tag_por_pontuacao(media_val)
                fill_map = {
                    "excelente": excelente_fill,
                    "bom": bom_fill,
                    "baixo": baixo_fill,
                    "texto": texto_fill,
                }
                fill = fill_map.get(tag)
                if fill and media_val > 0:
                    ws.cell(row=row_idx, column=col - 1).fill = fill

            # Ajustar larguras
            larguras = {"A": 10, "B": 22, "C": 30}
            for col_letter in ["A", "B", "C"]:
                ws.column_dimensions[col_letter].width = larguras.get(col_letter, 18)
            for i in range(4, 16):  # Jan-Dez
                ws.column_dimensions[get_column_letter(i)].width = 14
            ws.column_dimensions[get_column_letter(16)].width = 12  # MEDIA
            ws.column_dimensions[get_column_letter(17)].width = 22  # CONTRATO

            # Formato numerico para colunas de pontuacao e media — 2 casas decimais
            for row_idx in range(2, len(self.dados) + 2):
                for col_idx in range(4, 17):  # Jan-Dez + MEDIA
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            wb.save(caminho)
            messagebox.showinfo("Sucesso", f"Arquivo exportado:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar:\n{e}")

    # ---- RESETAR DADOS ----
    def _resetar_dados(self):
        resposta = messagebox.askyesno(
            "Resetar Dados",
            "Deseja restaurar os dados originais?\n\n"
            "Isso ira apagar todas as alteracoes feitas\n"
            "e importacoes anteriores.\n\n"
            "O arquivo dados_lojas.json sera recriado."
        )
        if resposta:
            if os.path.exists(ARQUIVO_DADOS):
                try:
                    os.remove(ARQUIVO_DADOS)
                except Exception:
                    pass
            # Limpar pasta de anexos
            for f in os.listdir(PASTA_ANEXOS):
                fp = os.path.join(PASTA_ANEXOS, f)
                try:
                    if os.path.isfile(fp):
                        os.remove(fp)
                except Exception:
                    pass
            self.dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(self.dados)
            self._popular_combos()
            self._atualizar_tabela()
            self._atualizar_grafico()
            self._atualizar_resumo()

    # ---- IMPORTAR EXCEL ----
    def _importar_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Erro", "openpyxl não está instalado.")
            return
        caminho = filedialog.askopenfilename(
            title="Importar planilha de dados",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not caminho:
            return
        try:
            self._importar_excel_arquivo(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar:\n{e}")

    def _importar_excel_arquivo(self, caminho):
        """Importa dados de planilha Excel, mapeando colunas automaticamente."""
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{e}")
            return

        # Mapear cabecalhos para colunas
        header_row = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                header_row[str(val).strip().upper()] = col_idx

        # Mapeamento flexivel de colunas
        col_map = {}
        # ANO
        for key in ["ANO", "YEAR"]:
            if key in header_row:
                col_map["ano"] = header_row[key]
                break
        # ESTADO
        for key in ["ESTADOS", "ESTADO", "STATE", "UF"]:
            if key in header_row:
                col_map["estado"] = header_row[key]
                break
        # LOJA
        for key in ["LOJAS", "LOJA", "STORE", "SHOP"]:
            if key in header_row:
                col_map["loja"] = header_row[key]
                break
        # CONTRATO
        for key in ["CONTRATO", "CONTRACT"]:
            if key in header_row:
                col_map["contrato"] = header_row[key]
                break
        # MESES
        mes_map_upper = {m.upper(): mk for mk, mnome in MESES for m in [mk.upper(), mnome.upper(), MESES_CURTOS[MESES.index((mk, mnome))].upper()]}
        for key, col_idx in header_row.items():
            if key in mes_map_upper:
                mk = mes_map_upper[key]
                col_map[f"pont_{mk}"] = col_idx

        # MEDIA
        for key in ["MEDIA", "MÉDIA", "AVERAGE", "AVG"]:
            if key in header_row:
                col_map["media"] = header_row[key]
                break

        # Verificar se tem pelo menos estado + loja + ano
        tem_minimo = all(k in col_map for k in ["ano", "estado", "loja"])
        if not tem_minimo:
            messagebox.showwarning(
                "Importar",
                "Colunas mínimas não encontradas (ANO, ESTADOS, LOJAS).\n"
                f"Encontradas: {list(col_map.keys())}"
            )
            return

        # Importar linhas
        importados = 0
        atualizados = 0
        for row_idx in range(2, ws.max_row + 1):
            ano_val = ws.cell(row=row_idx, column=col_map["ano"]).value
            if ano_val is None:
                continue
            try:
                ano = int(ano_val)
            except (ValueError, TypeError):
                continue

            estado = str(ws.cell(row=row_idx, column=col_map["estado"]).value or "").strip()
            loja = str(ws.cell(row=row_idx, column=col_map["loja"]).value or "").strip()
            if not estado or not loja:
                continue

            contrato = str(ws.cell(row=row_idx, column=col_map.get("contrato", -1)).value or "ASSAÍ - ATACADISTA").strip() if "contrato" in col_map else "ASSAÍ - ATACADISTA"

            # Buscar ou criar registro
            reg = None
            for r in self.dados:
                if (str(r.get("estado", "")).strip().upper() == estado.upper() and
                    str(r.get("loja", "")).strip().upper() == loja.upper() and
                    r.get("ano") == ano):
                    reg = r
                    break

            if reg:
                atualizados += 1
            else:
                reg = _reg(estado, loja, contrato, ano)
                self.dados.append(reg)
                importados += 1

            reg["contrato"] = contrato

            # Atualizar pontuacao mensal
            for mk, _ in MESES:
                campo_col = f"pont_{mk}"
                if campo_col in col_map:
                    cell_val = ws.cell(row=row_idx, column=col_map[campo_col]).value
                    if cell_val is not None:
                        if isinstance(cell_val, (int, float)):
                            reg[f"pont_{mk}"] = formatar_pontuacao(cell_val)
                        else:
                            val_str = str(cell_val).strip()
                            if val_str and eh_valor_texto(val_str):
                                reg[f"pont_{mk}"] = val_str
                            elif val_str:
                                try:
                                    f = float(val_str.replace(",", "."))
                                    reg[f"pont_{mk}"] = formatar_pontuacao(f)
                                except (ValueError, TypeError):
                                    reg[f"pont_{mk}"] = val_str

            calcular_media(reg)

        salvar_dados_json(self.dados)
        self._popular_combos()
        self._atualizar_tabela()
        self._atualizar_grafico()
        self._atualizar_resumo()

        messagebox.showinfo(
            "Importar",
            f"Importação concluída!\n\n"
            f"Novos registros: {importados}\n"
            f"Registros atualizados: {atualizados}"
        )

    # ---- RESUMO ----
    def _atualizar_resumo(self):
        n = len(self.dados)
        estados = len(obter_estados(self.dados))
        lojas = len(set(r.get("loja", "") for r in self.dados if r.get("loja")))
        self.title(f"{APP_TITLE} | {n} registros | {estados} estados | {lojas} lojas")


# ============================================================
# MAIN
# ============================================================
def main():
    try:
        app = RelatorioLojasApp()
        app.mainloop()
    except Exception as e:
        erro_txt = f"{type(e).__name__}: {e}"
        try:
            with open("erro_log.txt", "w", encoding="utf-8") as f:
                f.write(f"ERRO AO ABRIR O APLICATIVO\n")
                f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write(f"Python: {sys.version}\n")
                f.write(f"\n{erro_txt}\n")
        except Exception:
            pass
        try:
            import tkinter as _tk
            _root = _tk.Tk()
            _root.withdraw()
            _tk.messagebox.showerror(
                "Erro ao abrir o aplicativo",
                f"Ocorreu um erro e o app nao pode abrir:\n\n{e}\n\n"
                f"Detalhes salvos em: erro_log.txt"
            )
            _root.destroy()
        except Exception:
            pass
        print(erro_txt)


if __name__ == "__main__":
    main()
