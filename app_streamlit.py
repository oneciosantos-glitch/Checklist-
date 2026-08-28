"""Relatório de Lojas Assaí │ Sistema de Checklist / Pontuação
Versão Streamlit — Interface web moderna com gráfico + persistência JSON
Campos mensais: pontuação 0-100 ou texto (ADM FÉRIAS, FÉRIAS, SEM ADM)
MÉDIA = média aritmética dos valores numéricos mensais (textos excluídos)
Suporte a anexo de pontuação
"""

import streamlit as st
import json
import os
import sys
import re
import shutil
import base64
import hashlib
from datetime import datetime
from collections import defaultdict
import tempfile
import io
import copy

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patheffects
    import matplotlib.patches as mpatches
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    from scipy.interpolate import make_interp_spline
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# ============================================================
# CONFIGURACAO
# ============================================================
APP_TITLE = "Relatório de Lojas Assaí"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__)) if '__file__' in dir() else os.getcwd()
ARQUIVO_DADOS = os.path.join(SCRIPT_DIR, "dados_lojas.json")
PASTA_ANEXOS = os.path.join(SCRIPT_DIR, "anexos")

ANO_INICIAL = 2026  # Apenas anos a partir de 2026
VALORES_TEXTO = ["ADM FÉRIAS", "FÉRIAS", "SEM ADM"]
CAMPOS_CHAVE = ["estado", "loja", "contrato", "ano"]

MESES = [
    ("jan", "Janeiro"), ("fev", "Fevereiro"), ("mar", "Março"),
    ("abr", "Abril"), ("mai", "Maio"), ("jun", "Junho"),
    ("jul", "Julho"), ("ago", "Agosto"), ("set", "Setembro"),
    ("out", "Outubro"), ("nov", "Novembro"), ("dez", "Dezembro"),
]
MESES_CURTOS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
               "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

COLUNAS_EXPORTACAO = ["ANO", "ESTADOS", "LOJAS"] + MESES_CURTOS + ["MEDIA", "CONTRATO"]

# ============================================================
# MAPA DO BRASIL (coordenadas simplificadas dos 27 estados)
# ============================================================
MAPA_ESTADOS = {
    "AC": {"nome": "Acre", "coords": [[-73.608,-7.202],[-70.369,-8.141],[-66.627,-9.935],[-68.543,-11.109],[-70.622,-10.999],[-70.494,-9.426],[-71.212,-9.967],[-72.18,-10.0],[-72.357,-9.494],[-73.215,-9.411],[-72.937,-8.988],[-73.988,-7.555],[-73.608,-7.202]]},
    "AL": {"nome": "Alagoas", "coords": [[-36.952,-9.382],[-35.152,-8.913],[-35.301,-9.185],[-35.354,-9.255],[-36.391,-10.501],[-38.237,-9.329],[-37.76,-8.857],[-36.952,-9.382]]},
    "AM": {"nome": "Amazonas", "coords": [[-67.407,2.247],[-67.088,1.167],[-66.318,0.755],[-65.585,1.009],[-65.54,0.649],[-65.103,1.157],[-64.337,1.364],[-63.996,1.98],[-63.141,2.173],[-62.706,1.94],[-62.445,0.977],[-62.188,-0.33],[-62.51,-0.759],[-61.896,-1.395],[-61.474,-1.579],[-61.585,-0.937],[-61.216,-0.5],[-60.667,-0.894],[-60.309,-0.724],[-60.037,0.264],[-58.895,0.264],[-58.872,-0.343],[-58.323,-1.143],[-56.679,-2.212],[-56.098,-2.027],[-58.478,-6.699],[-58.136,-7.356],[-58.395,-8.78],[-61.985,-8.878],[-62.845,-7.986],[-63.62,-7.969],[-64.149,-8.959],[-64.839,-8.994],[-65.097,-9.432],[-65.246,-9.257],[-65.791,-9.585],[-66.408,-9.407],[-66.806,-9.814],[-70.369,-8.141],[-73.804,-7.111],[-73.137,-6.497],[-73.236,-6.031],[-72.814,-5.11],[-70.653,-4.127],[-69.964,-4.3],[-69.395,-1.132],[-70.057,-0.186],[-70.043,0.559],[-69.114,0.65],[-69.265,1.065],[-69.844,1.086],[-69.842,1.721],[-68.156,1.732],[-68.208,1.962],[-67.941,1.831],[-67.407,2.247]]},
    "AP": {"nome": "Amapá", "coords": [[-51.549,4.425],[-50.702,2.139],[-49.921,1.704],[-49.893,1.193],[-52.07,-1.236],[-52.933,-0.139],[-53.426,1.243],[-54.744,1.776],[-54.876,2.427],[-52.944,2.169],[-51.549,4.425]]},
    "BA": {"nome": "Bahia", "coords": [[-39.288,-8.563],[-38.296,-9.022],[-37.736,-10.332],[-38.229,-10.915],[-37.813,-11.514],[-37.343,-11.443],[-38.049,-12.634],[-38.965,-13.283],[-38.856,-15.86],[-39.135,-17.688],[-39.67,-18.349],[-40.623,-17.406],[-39.856,-16.113],[-40.23,-15.803],[-41.331,-15.744],[-41.8,-15.101],[-43.176,-14.65],[-43.531,-14.815],[-44.215,-14.233],[-46.077,-15.264],[-45.907,-14.353],[-46.265,-14.098],[-46.041,-13.28],[-46.315,-13.303],[-46.114,-12.918],[-46.397,-12.04],[-46.083,-11.636],[-46.617,-11.289],[-45.603,-10.108],[-45.248,-10.822],[-44.134,-10.636],[-43.662,-10.004],[-43.849,-9.548],[-43.485,-9.265],[-42.765,-9.616],[-41.113,-8.704],[-40.623,-9.482],[-39.288,-8.563]]},
    "CE": {"nome": "Ceará", "coords": [[-40.018,-2.837],[-37.252,-4.832],[-37.64,-4.926],[-38.578,-6.28],[-38.765,-6.994],[-38.534,-7.293],[-38.966,-7.845],[-39.662,-7.31],[-40.548,-7.392],[-40.37,-6.803],[-40.732,-6.654],[-40.925,-5.181],[-41.249,-4.869],[-41.322,-2.921],[-40.018,-2.837]]},
    "ES": {"nome": "Espírito Santo", "coords": [[-40.424,-20.635],[-40.957,-21.303],[-41.718,-21.123],[-41.858,-20.372],[-41.382,-20.188],[-40.949,-19.473],[-41.158,-18.308],[-40.527,-17.891],[-39.666,-18.332],[-39.689,-19.306],[-40.424,-20.635]]},
    "GO": {"nome": "Goiás", "coords": [[-50.158,-12.412],[-50.292,-12.839],[-49.369,-13.274],[-49.119,-12.79],[-48.586,-13.317],[-48.173,-13.148],[-47.679,-13.467],[-47.634,-13.104],[-47.427,-13.289],[-46.114,-12.918],[-46.265,-14.098],[-45.907,-14.353],[-46.088,-14.936],[-46.503,-14.704],[-46.502,-15.052],[-46.924,-15.058],[-46.812,-15.885],[-47.319,-16.036],[-47.417,-15.5],[-48.197,-15.501],[-48.279,-16.051],[-47.304,-16.06],[-47.458,-16.502],[-47.126,-16.98],[-47.541,-17.454],[-47.283,-18.058],[-47.954,-18.5],[-48.936,-18.306],[-49.378,-18.642],[-50.309,-18.698],[-50.842,-19.499],[-52.916,-18.639],[-52.758,-18.348],[-53.101,-18.31],[-53.246,-17.532],[-52.681,-16.303],[-51.086,-14.917],[-50.871,-13.733],[-50.158,-12.412]]},
    "MA": {"nome": "Maranhão", "coords": [[-47.031,-8.985],[-46.466,-8.066],[-47.043,-8.053],[-47.746,-7.201],[-47.378,-6.27],[-47.5,-5.525],[-48.363,-5.168],[-48.755,-5.349],[-47.088,-3.855],[-46.151,-1.224],[-45.846,-1.045],[-45.719,-1.404],[-45.58,-1.257],[-45.41,-1.289],[-45.488,-1.431],[-45.487,-1.53],[-45.443,-1.543],[-45.448,-1.449],[-45.317,-1.318],[-45.352,-1.736],[-45.101,-1.36],[-44.816,-1.418],[-44.642,-1.624],[-44.814,-1.815],[-44.593,-1.744],[-44.326,-2.5],[-43.615,-2.219],[-41.823,-2.719],[-42.989,-4.234],[-42.919,-6.67],[-44.033,-6.76],[-45.456,-7.67],[-45.994,-8.926],[-45.946,-10.258],[-46.367,-10.168],[-47.031,-8.985]]},
    "MG": {"nome": "Minas Gerais", "coords": [[-44.209,-14.244],[-43.783,-14.339],[-43.883,-14.653],[-43.531,-14.815],[-43.176,-14.65],[-41.8,-15.101],[-41.331,-15.744],[-40.707,-15.666],[-39.856,-16.113],[-40.57,-17.061],[-40.222,-17.98],[-40.882,-17.97],[-40.771,-18.155],[-41.158,-18.308],[-40.944,-19.46],[-41.382,-20.188],[-41.847,-20.329],[-42.271,-21.715],[-46.345,-22.904],[-46.723,-22.306],[-46.509,-21.469],[-47.011,-21.422],[-47.466,-19.964],[-48.823,-20.161],[-48.899,-20.441],[-49.551,-19.905],[-50.471,-19.779],[-51.0,-20.085],[-50.962,-19.484],[-50.309,-18.698],[-49.378,-18.642],[-48.936,-18.306],[-47.954,-18.5],[-47.283,-18.058],[-47.538,-17.388],[-47.126,-16.98],[-47.458,-16.502],[-47.3,-16.017],[-46.812,-15.885],[-46.918,-15.049],[-46.502,-15.052],[-46.474,-14.705],[-46.003,-14.902],[-46.119,-15.192],[-46.052,-15.259],[-44.209,-14.244]]},
    "MS": {"nome": "Mato Grosso do Sul", "coords": [[-53.874,-17.922],[-53.071,-18.039],[-53.069,-18.342],[-52.758,-18.348],[-52.916,-18.639],[-51.057,-19.329],[-51.001,-20.096],[-52.408,-22.141],[-53.607,-22.951],[-54.129,-23.982],[-55.347,-23.994],[-55.849,-22.284],[-57.991,-22.09],[-57.819,-20.942],[-58.167,-20.171],[-57.859,-19.97],[-58.131,-19.758],[-57.453,-18.231],[-57.795,-17.56],[-57.452,-17.902],[-56.113,-17.167],[-55.127,-17.652],[-54.302,-17.661],[-53.708,-17.228],[-53.874,-17.922]]},
    "MT": {"nome": "Mato Grosso", "coords": [[-60.716,-13.682],[-59.774,-12.341],[-60.108,-11.839],[-59.976,-11.122],[-61.55,-10.986],[-61.582,-8.798],[-58.415,-8.792],[-58.138,-7.349],[-57.592,-8.756],[-56.761,-9.405],[-50.224,-9.841],[-50.739,-11.435],[-50.502,-12.884],[-51.086,-14.917],[-52.681,-16.303],[-53.218,-17.299],[-53.071,-18.039],[-53.948,-17.923],[-53.708,-17.228],[-54.084,-17.619],[-55.127,-17.652],[-56.113,-17.167],[-57.452,-17.902],[-58.395,-17.184],[-58.321,-16.264],[-60.171,-16.265],[-60.564,-15.108],[-60.244,-15.096],[-60.381,-13.987],[-60.716,-13.682]]},
    "PA": {"nome": "Pará", "coords": [[-48.472,-0.499],[-46.428,-1.065],[-46.19,-0.894],[-46.272,-1.172],[-46.072,-1.019],[-46.28,-2.153],[-47.581,-4.52],[-48.755,-5.349],[-48.138,-5.602],[-48.231,-5.946],[-49.209,-6.925],[-49.215,-8.194],[-50.224,-9.841],[-56.754,-9.406],[-57.592,-8.756],[-58.478,-6.699],[-56.098,-2.027],[-56.679,-2.212],[-58.43,-1.027],[-58.872,-0.343],[-58.895,1.228],[-57.304,2.0],[-55.956,1.845],[-55.978,2.528],[-54.954,2.584],[-54.744,1.776],[-53.426,1.243],[-52.933,-0.139],[-52.1,-1.226],[-50.157,0.705],[-50.061,0.339],[-48.412,-0.257],[-48.472,-0.499]]},
    "PB": {"nome": "Paraíba", "coords": [[-37.227,-6.035],[-37.484,-6.71],[-36.718,-6.982],[-36.394,-6.294],[-34.971,-6.485],[-34.826,-7.547],[-35.479,-7.445],[-36.991,-8.303],[-37.355,-7.975],[-36.984,-7.482],[-37.233,-7.275],[-38.077,-7.83],[-38.593,-7.754],[-38.602,-6.389],[-38.115,-6.521],[-37.227,-6.035]]},
    "PE": {"nome": "Pernambuco", "coords": [[-37.177,-7.309],[-36.984,-7.482],[-37.355,-7.975],[-36.991,-8.303],[-35.479,-7.445],[-34.84,-7.543],[-35.152,-8.913],[-35.896,-8.854],[-36.952,-9.382],[-37.76,-8.857],[-38.237,-9.329],[-38.479,-8.85],[-39.383,-8.533],[-40.623,-9.482],[-40.921,-8.835],[-41.358,-8.707],[-40.589,-8.138],[-40.548,-7.392],[-39.662,-7.31],[-39.091,-7.858],[-38.715,-7.622],[-38.077,-7.83],[-37.177,-7.309]]},
    "PI": {"nome": "Piauí", "coords": [[-41.739,-2.806],[-41.257,-3.004],[-41.249,-4.869],[-40.925,-5.181],[-40.732,-6.654],[-40.37,-6.803],[-40.713,-7.473],[-40.589,-8.138],[-41.838,-9.242],[-42.765,-9.616],[-43.485,-9.265],[-43.849,-9.548],[-43.662,-10.004],[-44.13,-10.633],[-45.248,-10.822],[-45.579,-10.122],[-45.955,-10.218],[-45.994,-8.926],[-45.496,-7.75],[-44.053,-6.768],[-42.919,-6.67],[-42.989,-4.234],[-41.739,-2.806]]},
    "PR": {"nome": "Paraná", "coords": [[-52.972,-22.57],[-49.727,-23.108],[-49.305,-24.672],[-48.581,-24.671],[-48.556,-25.084],[-48.023,-25.23],[-48.59,-25.976],[-49.555,-26.237],[-50.571,-26.003],[-51.411,-26.717],[-53.551,-26.292],[-53.892,-25.622],[-54.593,-25.592],[-54.341,-24.129],[-53.607,-22.951],[-52.972,-22.57]]},
    "RJ": {"nome": "Rio de Janeiro", "coords": [[-44.668,-23.054],[-44.724,-23.368],[-44.875,-23.249],[-44.802,-22.999],[-44.161,-22.678],[-44.793,-22.387],[-42.271,-21.715],[-41.875,-20.766],[-41.718,-21.123],[-40.957,-21.303],[-40.985,-21.999],[-41.96,-22.534],[-42.013,-22.997],[-43.051,-22.982],[-43.085,-22.677],[-43.286,-23.016],[-43.553,-23.076],[-43.711,-23.056],[-43.856,-22.902],[-44.668,-23.054]]},
    "RN": {"nome": "Rio Grande do Norte", "coords": [[-37.64,-4.926],[-35.489,-5.158],[-34.969,-6.488],[-36.394,-6.294],[-36.718,-6.982],[-37.484,-6.71],[-37.174,-6.048],[-38.115,-6.521],[-38.577,-6.347],[-37.64,-4.926]]},
    "RO": {"nome": "Rondônia", "coords": [[-65.374,-9.699],[-66.806,-9.814],[-66.408,-9.407],[-65.097,-9.432],[-64.839,-8.994],[-64.143,-8.953],[-63.62,-7.969],[-62.866,-7.975],[-62.124,-8.801],[-61.468,-8.917],[-61.55,-10.986],[-59.986,-11.114],[-60.108,-11.839],[-59.774,-12.341],[-60.708,-13.692],[-61.84,-13.548],[-63.157,-12.613],[-64.291,-12.5],[-65.031,-11.995],[-65.361,-11.251],[-65.374,-9.699]]},
    "RR": {"nome": "Roraima", "coords": [[-59.916,3.146],[-59.751,1.862],[-58.886,1.261],[-58.895,0.264],[-60.037,0.264],[-60.531,-0.875],[-61.087,-0.5],[-61.428,-0.634],[-61.482,-1.58],[-62.51,-0.759],[-62.188,-0.33],[-62.706,1.94],[-64.055,2.498],[-64.185,3.56],[-64.824,4.244],[-63.964,3.868],[-63.206,3.952],[-62.96,3.608],[-62.747,4.035],[-60.996,4.518],[-60.591,4.927],[-60.723,5.22],[-60.21,5.271],[-59.99,4.987],[-60.162,4.508],[-59.675,4.373],[-59.517,3.943],[-59.916,3.146]]},
    "RS": {"nome": "Rio Grande do Sul", "coords": [[-51.095,-30.381],[-52.256,-31.849],[-52.098,-32.162],[-52.629,-33.116],[-53.416,-33.748],[-53.43,-33.156],[-53.123,-32.791],[-52.75,-32.862],[-52.622,-32.143],[-53.076,-32.656],[-53.388,-32.586],[-55.577,-30.833],[-56.009,-31.081],[-56.807,-30.104],[-57.643,-30.188],[-54.812,-27.529],[-53.874,-27.127],[-52.166,-27.273],[-50.625,-28.391],[-49.691,-28.618],[-50.166,-29.247],[-49.711,-29.325],[-50.769,-31.11],[-52.081,-32.157],[-52.098,-31.835],[-51.238,-31.458],[-50.574,-30.481],[-50.597,-30.194],[-50.93,-30.435],[-51.295,-30.001],[-51.095,-30.381]]},
    "SC": {"nome": "Santa Catarina", "coords": [[-53.834,-27.169],[-53.643,-26.252],[-53.281,-26.247],[-51.411,-26.717],[-50.571,-26.003],[-49.555,-26.237],[-48.643,-25.956],[-48.465,-27.145],[-48.616,-27.251],[-48.743,-28.508],[-50.065,-29.341],[-49.765,-28.46],[-50.625,-28.391],[-52.166,-27.273],[-53.834,-27.169]]},
    "SE": {"nome": "Sergipe", "coords": [[-37.96,-9.533],[-36.393,-10.498],[-37.518,-11.548],[-38.24,-10.876],[-37.736,-10.332],[-37.96,-9.533]]},
    "SP": {"nome": "São Paulo", "coords": [[-46.138,-23.859],[-48.099,-25.311],[-48.251,-24.978],[-48.582,-25.051],[-48.581,-24.671],[-49.305,-24.672],[-49.2,-24.344],[-49.986,-22.897],[-53.108,-22.597],[-52.408,-22.141],[-51.586,-20.634],[-50.576,-19.816],[-49.265,-19.962],[-48.899,-20.441],[-48.823,-20.161],[-47.473,-19.962],[-47.011,-21.422],[-46.509,-21.469],[-46.723,-22.306],[-46.359,-22.896],[-44.809,-22.405],[-44.162,-22.674],[-44.792,-22.982],[-44.724,-23.368],[-45.406,-23.623],[-45.405,-23.82],[-46.138,-23.859]]},
    "TO": {"nome": "Tocantins", "coords": [[-47.033,-8.982],[-46.367,-10.168],[-45.699,-10.166],[-46.617,-11.289],[-46.086,-11.622],[-46.397,-12.04],[-46.119,-12.925],[-47.427,-13.289],[-47.634,-13.104],[-47.679,-13.467],[-48.173,-13.148],[-48.586,-13.317],[-48.857,-12.805],[-49.369,-13.274],[-50.292,-12.839],[-50.142,-12.396],[-50.622,-12.819],[-50.603,-10.661],[-49.283,-8.379],[-49.209,-6.925],[-48.382,-6.379],[-48.132,-5.618],[-48.745,-5.369],[-48.519,-5.192],[-47.5,-5.525],[-47.378,-6.27],[-47.746,-7.201],[-47.043,-8.053],[-46.477,-8.012],[-47.033,-8.982]]},
    "DF": {"nome": "Distrito Federal", "coords": [[-47.417,-15.5],[-47.308,-16.05],[-48.279,-16.051],[-48.197,-15.501],[-47.417,-15.5]]},
}

ESTADO_PARA_SIGLA = {
    "CEARA": "CE", "BAHIA": "BA", "SERGIPE": "SE", "PERNAMBUCO": "PE",
    "PARAIBA": "PB", "RIO GRANDE DO NORTE": "RN", "PIAUI": "PI",
    "MARANHAO": "MA", "AMAPA": "AP", "MANAUS": "AM", "PARA": "PA", "RORAIMA": "RR",
}

# Criar pasta de anexos
os.makedirs(PASTA_ANEXOS, exist_ok=True)


# ============================================================
# DADOS REAIS
# ============================================================
def _reg(estado, loja, contrato, ano):
    reg = {"estado": estado, "loja": loja, "contrato": contrato, "ano": ano}
    for mk, _ in MESES:
        reg[f"pont_{mk}"] = ""
    reg["media"] = 0.0
    reg["anexo"] = ""
    return reg


DADOS_REAIS = [
    _reg("CEARA","ASSAÍ - CAUCAIA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - PARANGABA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MESSEJANA","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MISTER HALL","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - SOBRAL","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - WASHIGUNTON SOARES","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - BEZERRA DE MENEZES","ASSAÍ - ATACADISTA",2026),
    _reg("CEARA","ASSAÍ - MARACANAÚ","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - CAMASSARI","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - FEIRA DE SANTANA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - GUANAMBÍ","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - MUSSURUNGA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - VITÓRIA DA CONQUISTA","ASSAÍ - ATACADISTA",2026),
    _reg("BAHIA","ASSAÍ - JEQUIÉ","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - IMBIRIBEIRA","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - AV.RECIFE","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - CAMARAGIBE","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - CARUARU","ASSAÍ - ATACADISTA",2026),
    _reg("PERNAMBUCO","ASSAÍ - PETROLINA","ASSAÍ - ATACADISTA",2026),
    _reg("SERGIPE","ASSAÍ - ARACAJU","ASSAÍ - ATACADISTA",2026),
    _reg("SERGIPE","ASSAÍ - LAGARTO","ASSAÍ - ATACADISTA",2026),
    _reg("SERGIPE","ASSAÍ - ITABAIANA","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - BELÉM","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - ANANINDEUA","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - CASTANHAL","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ASSAÍ - MARABÁ","ASSAÍ - ATACADISTA",2026),
    _reg("PARA","ALGUSTO MONTE NEGRO","ASSAÍ - ATACADISTA",2026),
    _reg("MARANHAO","ASSAÍ - SÃO LUIS","ASSAÍ - ATACADISTA",2026),
    _reg("MARANHAO","ASSAÍ - IMPERATRIZ","ASSAÍ - ATACADISTA",2026),
    _reg("MARANHAO","ASSAÍ - TIMON","ASSAÍ - ATACADISTA",2026),
    _reg("PIAUI","ASSAÍ - TERESINA","ASSAÍ - ATACADISTA",2026),
    _reg("PIAUI","ASSAÍ - PIRIPIRI","ASSAÍ - ATACADISTA",2026),
    _reg("PIAUI","ASSAÍ - PICOS","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - NATAL","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - ARAPIRACA","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - MOSSORÓ","ASSAÍ - ATACADISTA",2026),
    _reg("RIO GRANDE DO NORTE","ASSAÍ - PARNAMIRIM","ASSAÍ - ATACADISTA",2026),
    _reg("PARAIBA","ASSAÍ - JOÃO PESSOA","ASSAÍ - ATACADISTA",2026),
    _reg("PARAIBA","ASSAÍ - CAMPINA GRANDE","ASSAÍ - ATACADISTA",2026),
    _reg("PARAIBA","ASSAÍ - SANTA RITA","ASSAÍ - ATACADISTA",2026),
    _reg("MANAUS","ASSAÍ - MANAUS","ASSAÍ - ATACADISTA",2026),
    _reg("MANAUS","ASSAÍ - MANAUS II","ASSAÍ - ATACADISTA",2026),
    _reg("AMAPA","ASSAÍ - MACAPÁ","ASSAÍ - ATACADISTA",2026),
    _reg("RORAIMA","ASSAÍ - BOA VISTA","ASSAÍ - ATACADISTA",2026),
]


# ============================================================
# FUNCOES UTILITARIAS
# ============================================================
def safe_float(v, default=0.0):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return default


def eh_valor_texto(val):
    if not isinstance(val, str):
        val = str(val).strip()
    return val.strip().upper() in [v.upper() for v in VALORES_TEXTO]


def calcular_media(registro):
    valores = []
    for mk, _ in MESES:
        val = str(registro.get(f"pont_{mk}", "")).strip()
        if val and not eh_valor_texto(val):
            try:
                f = float(val.replace(",", "."))
                valores.append(f)
            except (ValueError, TypeError):
                pass
    if valores:
        registro["media"] = round(sum(valores) / len(valores), 2)
    else:
        registro["media"] = 0.0
    return registro


def formatar_pontuacao(val):
    try:
        f = float(str(val).replace(",", "."))
        txt = f"{f:.2f}".replace(".", ",")
        return txt
    except (ValueError, TypeError):
        return str(val)


def formatar_media(val):
    try:
        f = float(str(val).replace(",", "."))
        txt = f"{f:.2f}".replace(".", ",")
        return txt
    except (ValueError, TypeError):
        return ""


def cor_por_pontuacao(pont):
    try:
        val = float(str(pont).replace(",", "."))
        if val >= 90:
            return "#16a34a"
        elif val >= 70:
            return "#d97706"
        else:
            return "#dc2626"
    except (ValueError, TypeError):
        return "#6b7280"


def tag_por_pontuacao(pont):
    try:
        val = float(str(pont).replace(",", "."))
        if val >= 90:
            return "excelente"
        elif val >= 70:
            return "bom"
        elif val > 0:
            return "baixo"
        else:
            return "vazio"
    except (ValueError, TypeError):
        return "texto"


def _ano_valido(r):
    """Retorna True se o registro tem ano >= ANO_INICIAL."""
    try:
        return int(r.get("ano", 0)) >= ANO_INICIAL
    except (ValueError, TypeError):
        return False


def obter_estados(dados):
    vistos = []
    for r in dados:
        e = str(r.get("estado", "")).strip()
        if e and e not in vistos:
            vistos.append(e)
    return vistos


def obter_lojas_por_estado(dados, estado):
    vistos = []
    for r in dados:
        if str(r.get("estado", "")).strip().upper() == estado.upper():
            l = str(r.get("loja", "")).strip()
            if l and l not in vistos:
                vistos.append(l)
    return vistos


def obter_anos_por_estado_loja(dados, estado, loja=None):
    anos = set()
    for r in dados:
        if str(r.get("estado", "")).strip().upper() != estado.upper():
            continue
        if loja and loja != "TODAS AS LOJAS" and str(r.get("loja", "")).strip().upper() != loja.upper():
            continue
        try:
            a = int(r.get("ano", 0))
            if a >= ANO_INICIAL:
                anos.add(a)
        except (ValueError, TypeError):
            pass
    return sorted(anos)


def buscar_registro(dados, estado, loja, ano):
    try:
        ano_int = int(ano)
    except (ValueError, TypeError):
        return None
    for r in dados:
        if (str(r.get("estado", "")).strip().upper() == estado.upper() and
            str(r.get("loja", "")).strip().upper() == loja.upper() and
            r.get("ano") == ano_int):
            return r
    return None


def _nome_arquivo_anexo(estado, loja, ano):
    safe = re.sub(r'[^\w\s\-\.]', '', f"{estado}_{loja}_{ano}")
    safe = re.sub(r'\s+', '_', safe.strip())
    return safe


def _normalizar_pontuacoes(reg):
    for mk, _ in MESES:
        campo = f"pont_{mk}"
        val = reg.get(campo, "")
        if val is None or val == "":
            reg[campo] = ""
        elif isinstance(val, (int, float)):
            if eh_valor_texto(str(val)):
                reg[campo] = str(val)
            else:
                try:
                    f = float(val)
                    reg[campo] = formatar_pontuacao(f)
                except (ValueError, TypeError):
                    reg[campo] = str(val)
        elif isinstance(val, str):
            val_s = val.strip()
            if not val_s:
                reg[campo] = ""
            elif eh_valor_texto(val_s):
                reg[campo] = val_s
            else:
                try:
                    f = float(val_s.replace(",", "."))
                    reg[campo] = formatar_pontuacao(f)
                except (ValueError, TypeError):
                    reg[campo] = val_s
    media_val = reg.get("media", 0)
    if isinstance(media_val, (int, float)):
        reg["media"] = round(float(media_val), 2)
    elif isinstance(media_val, str) and media_val.strip():
        try:
            reg["media"] = round(float(media_val.replace(",", ".")), 2)
        except (ValueError, TypeError):
            reg["media"] = 0.0
    return reg


def _migrar_dados_antigos(dados):
    convertidos = 0
    novos_dados = []
    for r in dados:
        novo = dict(r)
        if any(f"meta_{mk}" in r or f"realizado_{mk}" in r for mk, _ in MESES):
            for mk, _ in MESES:
                campo = f"pont_{mk}"
                if campo not in novo:
                    meta_v = safe_float(r.get(f"meta_{mk}", 0))
                    real_v = safe_float(r.get(f"realizado_{mk}", 0))
                    if meta_v > 0 and real_v > 0:
                        pct = (real_v / meta_v) * 100
                        novo[campo] = formatar_pontuacao(pct)
                    elif meta_v > 0:
                        novo[campo] = "0"
                    else:
                        novo[campo] = ""
                novo.pop(f"meta_{mk}", None)
                novo.pop(f"realizado_{mk}", None)
            convertidos += 1
        if "anexo" not in novo:
            novo["anexo"] = ""
        calcular_media(novo)
        _normalizar_pontuacoes(novo)
        novos_dados.append(novo)
    if convertidos > 0:
        print(f"Migrados {convertidos} registros do formato antigo.")
    return novos_dados


def migrar_registro(reg):
    for mk, _ in MESES:
        campo = f"pont_{mk}"
        if campo not in reg:
            reg[campo] = ""
    if "media" not in reg:
        calcular_media(reg)
    if "anexo" not in reg:
        reg["anexo"] = ""
    return reg


def carregar_dados():
    if os.path.exists(ARQUIVO_DADOS):
        try:
            with open(ARQUIVO_DADOS, "r", encoding="utf-8") as f:
                conteudo = f.read().strip()
            if not conteudo:
                raise ValueError("Arquivo vazio")
            dados = json.loads(conteudo)
            if not isinstance(dados, list):
                raise ValueError("Formato invalido")
            dados = _migrar_dados_antigos(dados)
            for r in dados:
                migrar_registro(r)
                _normalizar_pontuacoes(r)
            # Remover registros com ano anterior a ANO_INICIAL
            antes = len(dados)
            dados = [r for r in dados if _ano_valido(r)]
            if len(dados) < antes:
                salvar_dados_json(dados)  # Persistir limpeza de anos antigos
            return dados
        except (json.JSONDecodeError, ValueError) as e:
            print(f"JSON corrompido: {e}")
            try:
                bak = ARQUIVO_DADOS + ".corrompido_" + datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copy2(ARQUIVO_DADOS, bak)
                os.remove(ARQUIVO_DADOS)
            except Exception:
                pass
            dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(dados)
            return dados
        except Exception as e:
            print(f"Erro inesperado: {e}")
            try:
                os.remove(ARQUIVO_DADOS)
            except Exception:
                pass
            dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(dados)
            return dados
    dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
    salvar_dados_json(dados)
    return dados


def salvar_dados_json(dados):
    try:
        with open(ARQUIVO_DADOS, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Erro ao salvar JSON: {e}")


# ============================================================
# GRAFICOS COM MAPA DO BRASIL
# ============================================================
def desenhar_mapa_fundo(ax, estado_selecionado=""):
    """Desenha o mapa do Brasil como marca d'água no fundo do gráfico."""
    if not HAS_MATPLOTLIB:
        return

    sigla_destaque = ESTADO_PARA_SIGLA.get(estado_selecionado, None)

    LON_MIN, LON_MAX = -74, -34
    LAT_MIN, LAT_MAX = -34, 6

    MAP_LEFT = 0.03
    MAP_RIGHT = 0.97
    MAP_BOTTOM = 0.02
    MAP_TOP = 0.98

    lon_range = LON_MAX - LON_MIN
    lat_range = LAT_MAX - LAT_MIN
    aspect = lat_range / lon_range

    axes_w = MAP_RIGHT - MAP_LEFT
    axes_h = MAP_TOP - MAP_BOTTOM
    if aspect >= 1.0:
        map_h = axes_w * aspect
        if map_h > axes_h:
            map_h = axes_h
            map_w = map_h / aspect
        else:
            map_w = axes_w
    else:
        map_w = axes_h / aspect
        if map_w > axes_w:
            map_w = axes_w
            map_h = map_w * aspect
        else:
            map_h = axes_h

    offset_x = MAP_LEFT + (axes_w - map_w) / 2
    offset_y = MAP_BOTTOM + (axes_h - map_h) / 2

    def lon_to_ax(lon):
        return offset_x + (lon - LON_MIN) / (LON_MAX - LON_MIN) * map_w

    def lat_to_ax(lat):
        return offset_y + (lat - LAT_MIN) / (LAT_MAX - LAT_MIN) * map_h

    cor_destaque = (0.231, 0.510, 0.965, 0.12)
    cor_outros = (0.88, 0.91, 0.94, 0.10)
    cor_borda = (0.70, 0.75, 0.80, 0.20)

    trans = ax.transAxes
    for sigla, state in MAPA_ESTADOS.items():
        coords_orig = state["coords"]
        coords_conv = [(lon_to_ax(c[0]), lat_to_ax(c[1])) for c in coords_orig]
        is_destaque = (sigla == sigla_destaque)
        fc = cor_destaque if is_destaque else cor_outros
        poly = mpatches.Polygon(coords_conv, closed=True,
                               facecolor=fc, edgecolor=cor_borda,
                               linewidth=0.6, zorder=0,
                               transform=trans)
        ax.add_patch(poly)

    if sigla_destaque:
        for sigla, state in MAPA_ESTADOS.items():
            if sigla != sigla_destaque:
                continue
            coords_orig = state["coords"]
            cx_lon = sum(c[0] for c in coords_orig) / len(coords_orig)
            cy_lat = sum(c[1] for c in coords_orig) / len(coords_orig)
            OFFSETS_ROTULO = {"SE": (0.6, -0.2), "PB": (0.4, 0.6), "RN": (0.6, 1.0)}
            ox_lon, oy_lat = OFFSETS_ROTULO.get(sigla, (0, 0))
            lx = lon_to_ax(cx_lon + ox_lon)
            ly = lat_to_ax(cy_lat + oy_lat)
            ax.text(lx, ly, sigla, ha="center", va="center",
                    fontsize=14, fontweight="bold", color="#3b82f6", alpha=0.40,
                    transform=trans,
                    path_effects=[
                        matplotlib.patheffects.withStroke(
                            linewidth=3, foreground="#1e40af")
                    ], zorder=1)


def gerar_grafico_mensal(dados, estado, loja, ano, todas_lojas):
    """Gráfico de barras/linha mostrando pontuação mensal para o ano selecionado."""
    if todas_lojas:
        registros = [r for r in dados
                     if str(r.get("estado", "")).strip().upper() == estado.upper()
                     and r.get("ano") == ano]
    else:
        registros = [r for r in dados
                     if str(r.get("estado", "")).strip().upper() == estado.upper()
                     and str(r.get("loja", "")).strip().upper() == loja.upper()
                     and r.get("ano") == ano]

    if not registros:
        return None

    ponts_agregado = {}
    contagens = {}
    textos_agregado = {}
    for r in registros:
        for mk, _ in MESES:
            val = str(r.get(f"pont_{mk}", "")).strip()
            if val and eh_valor_texto(val):
                if mk not in textos_agregado:
                    textos_agregado[mk] = []
                textos_agregado[mk].append(val)
            elif val:
                try:
                    f = float(val.replace(",", "."))
                    ponts_agregado[mk] = ponts_agregado.get(mk, 0) + f
                    contagens[mk] = contagens.get(mk, 0) + 1
                except (ValueError, TypeError):
                    pass

    ponts_mensal = []
    for mk, _ in MESES:
        if mk in ponts_agregado and mk in contagens:
            ponts_mensal.append(ponts_agregado[mk] / contagens[mk])
        elif mk in ponts_agregado:
            ponts_mensal.append(ponts_agregado[mk])
        else:
            ponts_mensal.append(0)

    if not any(p > 0 for p in ponts_mensal):
        return None

    if todas_lojas:
        titulo = f"Média das Lojas — {estado} ({ano})"
    else:
        titulo = f"{loja} — {ano}"

    fig = Figure(figsize=(10, 5), dpi=100, facecolor="#f8fafc")
    ax = fig.add_subplot(111)

    x_pos = np.arange(12) if HAS_NUMPY else list(range(12))

    # Spline suave
    usar_spline = False
    pontos_validos = [(i, p) for i, p in enumerate(ponts_mensal) if p > 0]
    if HAS_SCIPY and HAS_NUMPY and len(pontos_validos) >= 3:
        try:
            xi = np.array([pv[0] for pv in pontos_validos])
            yi = np.array([pv[1] for pv in pontos_validos])
            grau = min(3, len(pontos_validos) - 1)
            spl = make_interp_spline(xi, yi, k=grau)
            x_smooth = np.linspace(xi.min(), xi.max(), 200)
            _ = spl(x_smooth[:3])
            usar_spline = True
        except Exception:
            usar_spline = False

    # Cor da linha baseada na média
    media_val = sum(ponts_mensal) / max(1, sum(1 for p in ponts_mensal if p > 0))
    if media_val >= 90:
        cor_linha = "#22c55e"
    elif media_val >= 70:
        cor_linha = "#3b82f6"
    else:
        cor_linha = "#ef4444"

    if usar_spline:
        ax.plot(x_smooth, spl(x_smooth), color=cor_linha, linewidth=2.5)
        ax.scatter(x_pos, ponts_mensal, color=cor_linha, s=70, zorder=5,
                   edgecolors="white", linewidths=1.5)
    else:
        ax.plot(x_pos, ponts_mensal, color=cor_linha, linewidth=2.5, marker="o",
                markersize=8, markeredgecolor="white", markeredgewidth=2)

    # Preencher área sob a curva
    if HAS_NUMPY:
        x_arr = np.array(x_pos)
        y_arr = np.array(ponts_mensal)
        ax.fill_between(x_arr, y_arr, alpha=0.08, color=cor_linha)

    # Anotações dos valores
    for i, p in enumerate(ponts_mensal):
        if p > 0:
            ax.annotate(formatar_pontuacao(p), (i, p), textcoords="offset points",
                        xytext=(0, 12), ha="center", fontsize=12, fontweight="bold",
                        color=cor_linha)
        mk = MESES[i][0]
        if mk in textos_agregado and textos_agregado[mk]:
            txt = textos_agregado[mk][0]
            ax.annotate(txt, (i, 2), ha="center", fontsize=9, color="#6b7280",
                        fontstyle="italic")

    ax.set_title(titulo, fontsize=14, fontweight="bold", pad=10)
    ax.set_xlabel("Mês", fontsize=12)
    ax.set_xticks(x_pos)
    ax.set_xticklabels(MESES_CURTOS, fontsize=10)
    ax.set_ylim(0, 110)
    ax.set_facecolor("#ffffff")
    desenhar_mapa_fundo(ax, estado)
    ax.set_xlim(-0.5, 11.5)
    ax.set_ylim(0, 110)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["top"].set_visible(False)
    ax.grid(False)

    fig.tight_layout()
    return fig


def gerar_grafico_anual(dados, estado, loja, todas_lojas):
    """Gráfico de tendência: MÉDIA por ano."""
    if todas_lojas:
        registros = [r for r in dados
                     if str(r.get("estado", "")).strip().upper() == estado.upper()]
    else:
        registros = [r for r in dados
                     if str(r.get("estado", "")).strip().upper() == estado.upper()
                     and str(r.get("loja", "")).strip().upper() == loja.upper()]

    if not registros:
        return None

    soma_media = defaultdict(float)
    contagem = defaultdict(int)
    for r in registros:
        a = int(r.get("ano", 0))
        m = safe_float(r.get("media", 0))
        if m > 0:
            soma_media[a] += m
            contagem[a] += 1

    anos = sorted(soma_media.keys())
    medias = [soma_media[a] / contagem[a] if contagem[a] > 0 else 0 for a in anos]

    if not any(m > 0 for m in medias):
        return None

    if todas_lojas:
        titulo = f"Tendência Anual — {estado} (Todas as Lojas)"
    else:
        titulo = f"Tendência Anual — {loja}"

    fig = Figure(figsize=(10, 5), dpi=100, facecolor="#f8fafc")
    ax = fig.add_subplot(111)

    if HAS_NUMPY:
        x = np.array(anos)
        y = np.array(medias)
    else:
        x = anos
        y = medias

    # Spline suave
    usar_spline = False
    if HAS_SCIPY and HAS_NUMPY and len(anos) >= 3:
        anos_unicos = len(set(anos))
        if anos_unicos == len(anos):
            try:
                x_smooth = np.linspace(x.min(), x.max(), 200)
                grau = min(3, len(anos) - 1)
                spl = make_interp_spline(x, y, k=grau)
                _ = spl(x_smooth[:3])
                usar_spline = True
            except Exception:
                usar_spline = False

    # Cor da linha
    media_anual = sum(medias) / max(1, len([m for m in medias if m > 0]))
    if media_anual >= 90:
        cor_linha = "#22c55e"
    elif media_anual >= 70:
        cor_linha = "#3b82f6"
    else:
        cor_linha = "#ef4444"

    if usar_spline:
        ax.plot(x_smooth, spl(x_smooth), color=cor_linha, linewidth=2.5)
        ax.scatter(x, y, color=cor_linha, s=70, zorder=5, edgecolors="white", linewidths=1.5)
    else:
        ax.plot(x, y, color=cor_linha, linewidth=2.5, marker="o", markersize=8,
                markeredgecolor="white", markeredgewidth=2)

    # Anotações
    for a_val, m_val in zip(anos, y):
        if m_val > 0:
            ax.annotate(formatar_media(m_val), (a_val, m_val), textcoords="offset points",
                        xytext=(0, 14), ha="center", fontsize=12, color=cor_linha,
                        fontweight="bold")

    # Preencher área
    if HAS_NUMPY and len(anos) >= 2:
        ax.fill_between(x, y, alpha=0.08, color=cor_linha)

    ax.set_title(titulo, fontsize=14, fontweight="bold", pad=10)
    ax.set_xlabel("Ano", fontsize=12)
    ax.set_ylim(0, 110)
    ax.set_facecolor("#ffffff")
    desenhar_mapa_fundo(ax, estado)
    ax.set_xlim(min(anos) - 0.5, max(anos) + 0.5)
    ax.set_ylim(0, 110)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["top"].set_visible(False)
    ax.grid(False)

    fig.tight_layout()
    return fig


# ============================================================
# EXPORTAR EXCEL
# ============================================================
def exportar_excel_bytes(dados):
    """Gera arquivo Excel em memória e retorna bytes."""
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist Pontuação"

    cabecalho = COLUNAS_EXPORTACAO
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2d2d2d", end_color="2d2d2d", fill_type="solid")
    for col_idx, h in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    excelente_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    bom_fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
    baixo_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    texto_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")

    for row_idx, r in enumerate(dados, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=r.get("ano", "")); col += 1
        ws.cell(row=row_idx, column=col, value=r.get("estado", "")); col += 1
        ws.cell(row=row_idx, column=col, value=r.get("loja", "")); col += 1

        for mk, _ in MESES:
            val = str(r.get(f"pont_{mk}", "")).strip()
            if val and not eh_valor_texto(val):
                try:
                    ws.cell(row=row_idx, column=col, value=float(val.replace(",", ".")))
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col, value=val)
            elif val:
                ws.cell(row=row_idx, column=col, value=val)
            else:
                ws.cell(row=row_idx, column=col, value="")
            col += 1

        media_val = safe_float(r.get("media", 0))
        ws.cell(row=row_idx, column=col, value=media_val if media_val > 0 else "")
        col += 1

        ws.cell(row=row_idx, column=col, value=r.get("contrato", ""))

        tag = tag_por_pontuacao(media_val)
        fill_map = {"excelente": excelente_fill, "bom": bom_fill, "baixo": baixo_fill, "texto": texto_fill}
        fill = fill_map.get(tag)
        if fill and media_val > 0:
            ws.cell(row=row_idx, column=col - 1).fill = fill

    # Ajustar larguras
    larguras = {"A": 10, "B": 22, "C": 30}
    for col_letter in ["A", "B", "C"]:
        ws.column_dimensions[col_letter].width = larguras.get(col_letter, 18)
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(16)].width = 12
    ws.column_dimensions[get_column_letter(17)].width = 22

    # Formato numérico
    for row_idx in range(2, len(dados) + 2):
        for col_idx in range(4, 17):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# IMPORTAR EXCEL
# ============================================================
def importar_excel_arquivo(dados, caminho_ou_bytes):
    """Importa dados de planilha Excel. Retorna (dados_atualizados, importados, atualizados)."""
    if not HAS_OPENPYXL:
        return dados, 0, 0

    try:
        if isinstance(caminho_ou_bytes, (bytes, io.BytesIO)):
            if isinstance(caminho_ou_bytes, bytes):
                wb = openpyxl.load_workbook(io.BytesIO(caminho_ou_bytes), data_only=True)
            else:
                wb = openpyxl.load_workbook(caminho_ou_bytes, data_only=True)
        else:
            wb = openpyxl.load_workbook(caminho_ou_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        st.error(f"Não foi possível abrir o arquivo: {e}")
        return dados, 0, 0

    header_row = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_row[str(val).strip().upper()] = col_idx

    col_map = {}
    for key in ["ANO", "YEAR"]:
        if key in header_row:
            col_map["ano"] = header_row[key]
            break
    for key in ["ESTADOS", "ESTADO", "STATE", "UF"]:
        if key in header_row:
            col_map["estado"] = header_row[key]
            break
    for key in ["LOJAS", "LOJA", "STORE", "SHOP"]:
        if key in header_row:
            col_map["loja"] = header_row[key]
            break
    for key in ["CONTRATO", "CONTRACT"]:
        if key in header_row:
            col_map["contrato"] = header_row[key]
            break

    mes_map_upper = {m.upper(): mk for mk, mnome in MESES for m in [mk.upper(), mnome.upper(), MESES_CURTOS[MESES.index((mk, mnome))].upper()]}
    for key, col_idx in header_row.items():
        if key in mes_map_upper:
            mk = mes_map_upper[key]
            col_map[f"pont_{mk}"] = col_idx

    for key in ["MEDIA", "MÉDIA", "AVERAGE", "AVG"]:
        if key in header_row:
            col_map["media"] = header_row[key]
            break

    if not all(k in col_map for k in ["ano", "estado", "loja"]):
        st.warning(f"Colunas mínimas não encontradas (ANO, ESTADOS, LOJAS). Encontradas: {list(col_map.keys())}")
        return dados, 0, 0

    importados = 0
    atualizados = 0
    for row_idx in range(2, ws.max_row + 1):
        ano_val = ws.cell(row=row_idx, column=col_map["ano"]).value
        if ano_val is None:
            continue
        try:
            ano = int(ano_val)
        except (ValueError, TypeError):
            continue
        if ano < ANO_INICIAL:
            continue

        estado = str(ws.cell(row=row_idx, column=col_map["estado"]).value or "").strip()
        loja = str(ws.cell(row=row_idx, column=col_map["loja"]).value or "").strip()
        if not estado or not loja:
            continue

        contrato = str(ws.cell(row=row_idx, column=col_map.get("contrato", -1)).value or "ASSAÍ - ATACADISTA").strip() if "contrato" in col_map else "ASSAÍ - ATACADISTA"

        reg = None
        for r in dados:
            if (str(r.get("estado", "")).strip().upper() == estado.upper() and
                str(r.get("loja", "")).strip().upper() == loja.upper() and
                r.get("ano") == ano):
                reg = r
                break

        if reg:
            atualizados += 1
        else:
            reg = _reg(estado, loja, contrato, ano)
            dados.append(reg)
            importados += 1

        reg["contrato"] = contrato

        for mk, _ in MESES:
            campo_col = f"pont_{mk}"
            if campo_col in col_map:
                cell_val = ws.cell(row=row_idx, column=col_map[campo_col]).value
                if cell_val is not None:
                    if isinstance(cell_val, (int, float)):
                        reg[f"pont_{mk}"] = formatar_pontuacao(cell_val)
                    else:
                        val_str = str(cell_val).strip()
                        if val_str and eh_valor_texto(val_str):
                            reg[f"pont_{mk}"] = val_str
                        elif val_str:
                            try:
                                f = float(val_str.replace(",", "."))
                                reg[f"pont_{mk}"] = formatar_pontuacao(f)
                            except (ValueError, TypeError):
                                reg[f"pont_{mk}"] = val_str

        calcular_media(reg)

    return dados, importados, atualizados


# ============================================================
# APP STREAMLIT
# ============================================================
def main():
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="🏪",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # ---- Carregar dados na session_state ----
    if "dados" not in st.session_state:
        st.session_state.dados = carregar_dados()
    if "save_counter" not in st.session_state:
        st.session_state.save_counter = 0

    dados = st.session_state.dados

    # ---- Header ----
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d2d2d 100%);
        padding: 18px 30px;
        border-radius: 12px;
        margin-bottom: 20px;
    }
    .main-header h1 {
        color: #ffffff;
        font-size: 26px;
        font-weight: 700;
        margin: 0;
    }
    .main-header .subtitle {
        color: #a0a0a0;
        font-size: 14px;
        margin-top: 4px;
    }
    .media-box {
        padding: 12px 20px;
        border-radius: 10px;
        text-align: center;
        margin: 8px 0;
    }
    .media-box .value {
        font-size: 32px;
        font-weight: 800;
    }
    .media-box .label {
        font-size: 14px;
        color: #6b7280;
    }
    div[data-testid="stSidebar"] > div:first-child {
        background: linear-gradient(180deg, #f0f4ff 0%, #ffffff 100%);
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 24px;
        font-weight: 600;
    }
    </style>
    """, unsafe_allow_html=True)

    n_reg = len(dados)
    n_est = len(obter_estados(dados))
    n_lojas = len(set(r.get("loja", "") for r in dados if r.get("loja")))

    st.markdown(f"""
    <div class="main-header">
        <h1>🏪 Checklist de Pontuação │ Lojas Assaí</h1>
        <div class="subtitle">{n_reg} registros │ {n_est} estados │ {n_lojas} lojas</div>
    </div>
    """, unsafe_allow_html=True)

    # ---- SIDEBAR: Filtros ----
    with st.sidebar:
        st.markdown("### 🔍 Filtros")

        estados = obter_estados(dados)
        estado_sel = st.selectbox("Estado", options=estados, index=0 if estados else None)

        lojas = obter_lojas_por_estado(dados, estado_sel) if estado_sel else []
        lojas_options = ["TODAS AS LOJAS"] + lojas
        loja_sel = st.selectbox("Loja", options=lojas_options, index=0)
        todas_lojas = (loja_sel == "TODAS AS LOJAS")

        anos = obter_anos_por_estado_loja(dados, estado_sel, loja_sel if not todas_lojas else None)
        anos_str = [str(a) for a in anos]
        ano_sel = st.selectbox("Ano", options=anos_str, index=len(anos_str)-1 if anos_str else None)

        st.divider()

        # ---- Ações ----
        st.markdown("### ⚡ Ações")

        # Novo estado
        with st.expander("➕ Novo Estado"):
            novo_estado = st.text_input("Nome do novo Estado", key="novo_estado_input")
            if st.button("Registrar Estado", key="btn_novo_estado"):
                if novo_estado.strip():
                    nome = novo_estado.strip().upper()
                    if nome not in obter_estados(dados):
                        nova_loja_padrao = f"ASSAÍ - {nome}"
                        novo_reg = _reg(nome, nova_loja_padrao, "ASSAÍ - ATACADISTA", datetime.now().year)
                        calcular_media(novo_reg)
                        dados.append(novo_reg)
                        salvar_dados_json(dados)
                        st.session_state.dados = dados
                        st.success(f"Estado '{nome}' registrado com loja padrão!")
                        st.rerun()
                    else:
                        st.info(f"Estado '{nome}' já existe.")

        # Nova loja
        with st.expander("➕ Nova Loja"):
            nova_loja = st.text_input(f"Nome da nova loja em {estado_sel}", key="nova_loja_input")
            if st.button("Registrar Loja", key="btn_nova_loja"):
                if nova_loja.strip() and estado_sel:
                    nome = nova_loja.strip().upper()
                    lojas_atuais = obter_lojas_por_estado(dados, estado_sel)
                    if nome not in lojas_atuais:
                        ano_nova = datetime.now().year
                        novo_reg = _reg(estado_sel, nome, "ASSAÍ - ATACADISTA", ano_nova)
                        calcular_media(novo_reg)
                        dados.append(novo_reg)
                        salvar_dados_json(dados)
                        st.session_state.dados = dados
                        st.success(f"Loja '{nome}' registrada!")
                        st.rerun()
                    else:
                        st.info(f"Loja '{nome}' já existe.")

        # Novo ano
        with st.expander("➕ Novo Ano"):
            novo_ano = st.text_input(f"Novo ano (a partir de {ANO_INICIAL})", key="novo_ano_input")
            if st.button("Registrar Ano", key="btn_novo_ano"):
                if novo_ano.strip() and estado_sel and not todas_lojas:
                    try:
                        ano_int = int(novo_ano.strip())
                        if ANO_INICIAL <= ano_int <= 2100:
                            reg_existente = buscar_registro(dados, estado_sel, loja_sel, ano_int)
                            if not reg_existente:
                                novo_reg = _reg(estado_sel, loja_sel, "ASSAÍ - ATACADISTA", ano_int)
                                calcular_media(novo_reg)
                                dados.append(novo_reg)
                                salvar_dados_json(dados)
                                st.session_state.dados = dados
                                st.success(f"Ano {ano_int} registrado!")
                                st.rerun()
                            else:
                                st.info("Este registro já existe.")
                        else:
                            st.warning(f"Ano inválido. Use a partir de {ANO_INICIAL}.")
                    except ValueError:
                        st.warning("Digite um ano válido.")

        st.divider()

        # ---- Exportar / Importar ----
        st.markdown("### 📊 Excel")

        if HAS_OPENPYXL:
            excel_bytes = exportar_excel_bytes(dados)
            if excel_bytes:
                filename = f"checklist_assai_{datetime.now().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="📥 Exportar Excel",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_exportar"
                )

            uploaded_file = st.file_uploader(
                "📤 Importar Excel", type=["xlsx", "xls"],
                key="file_uploader_excel"
            )
            if uploaded_file is not None:
                file_bytes = uploaded_file.read()
                dados, imp, atl = importar_excel_arquivo(dados, file_bytes)
                if imp > 0 or atl > 0:
                    salvar_dados_json(dados)
                    st.session_state.dados = dados
                    st.success(f"Importação concluída! Novos: {imp}, Atualizados: {atl}")
                    st.rerun()

        st.divider()

        # ---- Resetar dados ----
        if st.button("🔄 Resetar Dados", key="btn_resetar", type="secondary"):
            if os.path.exists(ARQUIVO_DADOS):
                try:
                    os.remove(ARQUIVO_DADOS)
                except Exception:
                    pass
            # Limpar pasta de anexos
            for f_name in os.listdir(PASTA_ANEXOS):
                fp = os.path.join(PASTA_ANEXOS, f_name)
                try:
                    if os.path.isfile(fp):
                        os.remove(fp)
                except Exception:
                    pass
            st.session_state.dados = [calcular_media(dict(r)) for r in DADOS_REAIS]
            salvar_dados_json(st.session_state.dados)
            st.success("Dados restaurados aos valores originais!")
            st.rerun()

    # ---- MAIN AREA ----
    tab_pesquisa, tab_tabela = st.tabs(["📝 Pesquisa e Edição", "📋 Tabela de Dados"])

    # ============================================================
    # TAB: PESQUISA E EDIÇÃO
    # ============================================================
    with tab_pesquisa:
        col_edicao, col_grafico = st.columns([2, 3])

        with col_edicao:
            # ---- Registro selecionado ----
            if not todas_lojas and estado_sel and loja_sel and ano_sel:
                reg = buscar_registro(dados, estado_sel, loja_sel, ano_sel)

                # Contrato
                contrato_val = reg.get("contrato", "ASSAÍ - ATACADISTA") if reg else "ASSAÍ - ATACADISTA"
                contrato = st.text_input("Contrato", value=contrato_val, key="contrato_input")

                # Anexo
                st.markdown("#### 📎 Anexo da Pontuação")
                anexo_atual = reg.get("anexo", "") if reg else ""
                tem_anexo = bool(anexo_atual and os.path.exists(anexo_atual))

                if tem_anexo:
                    nome_arq = os.path.basename(anexo_atual)
                    st.info(f"📎 Anexo: **{nome_arq}**")
                    # Download do anexo
                    try:
                        with open(anexo_atual, "rb") as f:
                            anexo_bytes = f.read()
                        st.download_button(
                            label="⬇️ Baixar Anexo",
                            data=anexo_bytes,
                            file_name=nome_arq,
                            key="btn_baixar_anexo"
                        )
                    except Exception:
                        st.warning("Erro ao ler anexo.")

                    if st.button("🗑️ Remover Anexo", key="btn_remover_anexo"):
                        if os.path.exists(anexo_atual):
                            try:
                                os.remove(anexo_atual)
                            except Exception:
                                pass
                        if reg:
                            reg["anexo"] = ""
                            salvar_dados_json(dados)
                            st.session_state.dados = dados
                            st.rerun()
                else:
                    st.caption("Nenhum anexo")

                # Upload de anexo
                anexo_upload = st.file_uploader(
                    "Anexar arquivo", type=["png", "jpg", "jpeg", "bmp", "gif", "pdf", "xlsx", "csv", "docx", "txt"],
                    key="anexo_uploader"
                )
                if anexo_upload is not None:
                    anexo_bytes_data = anexo_upload.read()
                    nome_base = _nome_arquivo_anexo(estado_sel, loja_sel, ano_sel)
                    ext = os.path.splitext(anexo_upload.name)[1]
                    nome_destino = f"{nome_base}{ext}"
                    caminho_dest = os.path.join(PASTA_ANEXOS, nome_destino)
                    # Evitar conflito
                    if os.path.exists(caminho_dest):
                        ts = datetime.now().strftime("%H%M%S")
                        nome_destino = f"{nome_base}_{ts}{ext}"
                        caminho_dest = os.path.join(PASTA_ANEXOS, nome_destino)
                    try:
                        with open(caminho_dest, "wb") as f:
                            f.write(anexo_bytes_data)
                        if reg:
                            anexo_antigo = reg.get("anexo", "")
                            if anexo_antigo and anexo_antigo != caminho_dest and os.path.exists(anexo_antigo):
                                try:
                                    os.remove(anexo_antigo)
                                except Exception:
                                    pass
                            reg["anexo"] = caminho_dest
                        else:
                            try:
                                ano_int = int(ano_sel)
                            except (ValueError, TypeError):
                                ano_int = ANO_INICIAL
                            novo = _reg(estado_sel, loja_sel, contrato, ano_int)
                            novo["anexo"] = caminho_dest
                            calcular_media(novo)
                            dados.append(novo)
                        salvar_dados_json(dados)
                        st.session_state.dados = dados
                        st.success("Anexo salvo com sucesso!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao salvar anexo: {e}")

                # ---- MÉDIA ----
                if reg:
                    media_val = safe_float(reg.get("media", 0))
                    cor_media = cor_por_pontuacao(media_val)
                    if media_val > 0:
                        st.markdown(f"""
                        <div class="media-box" style="background: {cor_media}22; border: 2px solid {cor_media};">
                            <div class="label">MÉDIA</div>
                            <div class="value" style="color: {cor_media};">{formatar_media(media_val)}</div>
                        </div>
                        """, unsafe_allow_html=True)

                    # Contar meses
                    num_meses = 0
                    txt_meses = 0
                    for mk, _ in MESES:
                        val = str(reg.get(f"pont_{mk}", "")).strip()
                        if val:
                            if eh_valor_texto(val):
                                txt_meses += 1
                            else:
                                num_meses += 1
                    st.caption(f"{num_meses} meses com pontuação, {txt_meses} com status textual")
                else:
                    st.markdown("""
                    <div class="media-box" style="background: #f3f4f6; border: 2px solid #d1d5db;">
                        <div class="label">MÉDIA</div>
                        <div class="value" style="color: #6b7280;">—</div>
                    </div>
                    <p style='text-align:center; color:#6b7280; font-size:12px;'>Registro novo · sem dados</p>
                    """, unsafe_allow_html=True)

                # ---- Pontuação Mensal ----
                st.markdown("""
                <div style='padding:8px 0; border-bottom:1px solid #e2e8f0;'>
                    <b style='font-size:15px;'>📅 Pontuação Mensal</b><br>
                    <span style='font-size:11px; color:#6b7280;'>Use número (0-100) ou selecione: ADM FÉRIAS, FÉRIAS, SEM ADM</span>
                </div>
                """, unsafe_allow_html=True)

                pont_mensal = {}
                for mk, mnome in MESES:
                    if reg:
                        val = str(reg.get(f"pont_{mk}", "")).strip()
                    else:
                        val = ""

                    col1, col2, col3 = st.columns([1, 2, 2])
                    with col1:
                        st.markdown(f"**{mnome}**")
                    with col2:
                        if val and eh_valor_texto(val):
                            # Texto de status — mostrar como disabled + selectbox
                            st.text_input(f"Pont. {mk}", value=val, disabled=True, key=f"pont_{mk}", label_visibility="collapsed")
                            pont_mensal[mk] = val
                        else:
                            # Campo numérico
                            display_val = ""
                            if val:
                                try:
                                    f = float(val.replace(",", "."))
                                    display_val = formatar_pontuacao(f)
                                except (ValueError, TypeError):
                                    display_val = val
                            input_val = st.text_input(
                                f"Pont. {mk}", value=display_val,
                                placeholder="Pontuação",
                                key=f"pont_{mk}",
                                label_visibility="collapsed"
                            )
                            if input_val.strip():
                                if eh_valor_texto(input_val.strip()):
                                    pont_mensal[mk] = input_val.strip()
                                else:
                                    try:
                                        f = float(input_val.strip().replace(",", "."))
                                        if 0 <= f <= 100:
                                            pont_mensal[mk] = formatar_pontuacao(f)
                                        else:
                                            pont_mensal[mk] = input_val.strip()
                                    except (ValueError, TypeError):
                                        pont_mensal[mk] = input_val.strip()
                            else:
                                pont_mensal[mk] = ""
                    with col3:
                        status_opts = ["(número)"] + VALORES_TEXTO
                        current_status = val if val and eh_valor_texto(val) else "(número)"
                        status_sel = st.selectbox(
                            f"Status {mk}", options=status_opts,
                            index=status_opts.index(current_status) if current_status in status_opts else 0,
                            key=f"status_{mk}",
                            label_visibility="collapsed"
                        )
                        if status_sel != "(número)":
                            pont_mensal[mk] = status_sel

                # ---- AUTO-SAVE: persistir alterações automaticamente ----
                if reg:
                    alterado = False
                    reg["contrato"] = contrato.strip() or "ASSAÍ - ATACADISTA"
                    for mk, _ in MESES:
                        novo_val = pont_mensal.get(mk, "")
                        antigo_val = str(reg.get(f"pont_{mk}", "")).strip()
                        if novo_val != antigo_val:
                            reg[f"pont_{mk}"] = novo_val
                            alterado = True
                    if alterado:
                        calcular_media(reg)
                        salvar_dados_json(dados)
                        st.session_state.dados = dados

                st.divider()

                # ---- Botões de ação ----
                col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)

                with col_btn1:
                    if st.button("💾 Salvar", key="btn_salvar", type="primary", use_container_width=True):
                        if reg:
                            reg["contrato"] = contrato.strip() or "ASSAÍ - ATACADISTA"
                            for mk, _ in MESES:
                                reg[f"pont_{mk}"] = pont_mensal.get(mk, "")
                            calcular_media(reg)
                        else:
                            try:
                                ano_int = int(ano_sel)
                            except (ValueError, TypeError):
                                ano_int = ANO_INICIAL
                            novo = {"estado": estado_sel, "loja": loja_sel, "contrato": contrato.strip() or "ASSAÍ - ATACADISTA", "ano": ano_int}
                            for mk, _ in MESES:
                                novo[f"pont_{mk}"] = pont_mensal.get(mk, "")
                            novo["anexo"] = ""
                            calcular_media(novo)
                            dados.append(novo)
                        salvar_dados_json(dados)
                        st.session_state.dados = dados
                        st.success("✅ Dados salvos com sucesso!")
                        st.rerun()

                with col_btn2:
                    if st.button("🗑️ Excluir", key="btn_excluir", type="secondary", use_container_width=True):
                        if reg:
                            anexo = reg.get("anexo", "")
                            if anexo and os.path.exists(anexo):
                                try:
                                    os.remove(anexo)
                                except Exception:
                                    pass
                            dados.remove(reg)
                            salvar_dados_json(dados)
                            st.session_state.dados = dados
                            st.success("Registro excluído!")
                            st.rerun()

                with col_btn3:
                    if st.button("📄 Novo", key="btn_novo_registro", type="secondary", use_container_width=True):
                        if not reg:
                            try:
                                ano_int = int(ano_sel)
                            except (ValueError, TypeError):
                                ano_int = ANO_INICIAL
                            novo = _reg(estado_sel, loja_sel, "ASSAÍ - ATACADISTA", ano_int)
                            calcular_media(novo)
                            dados.append(novo)
                            salvar_dados_json(dados)
                            st.session_state.dados = dados
                            st.success("Novo registro criado!")
                            st.rerun()
                        else:
                            st.info("Este registro já existe. Edite os campos e clique Salvar.")

                with col_btn4:
                    if st.button("🧹 Limpar", key="btn_limpar", type="secondary", use_container_width=True):
                        if reg:
                            for mk, _ in MESES:
                                reg[f"pont_{mk}"] = ""
                            calcular_media(reg)
                            salvar_dados_json(dados)
                            st.session_state.dados = dados
                            st.success("Campos mensais limpos!")
                            st.rerun()

            else:
                # TODAS AS LOJAS selecionado
                st.info("Selecione uma loja específica para editar pontuações mensais.")

                # Mostrar MÉDIA agregada quando "TODAS AS LOJAS"
                if todas_lojas and estado_sel and ano_sel:
                    try:
                        ano = int(ano_sel)
                        regs_estado_ano = [r for r in dados
                                           if str(r.get("estado", "")).strip().upper() == estado_sel.upper()
                                           and r.get("ano") == ano]
                        medias = []
                        for r in regs_estado_ano:
                            m = safe_float(r.get("media", 0))
                            if m > 0:
                                medias.append(m)
                        if medias:
                            media_geral = sum(medias) / len(medias)
                            cor = cor_por_pontuacao(media_geral)
                            st.markdown(f"""
                            <div class="media-box" style="background: {cor}22; border: 2px solid {cor};">
                                <div class="label">MÉDIA GERAL ({estado_sel} — {ano_sel})</div>
                                <div class="value" style="color: {cor};">{formatar_media(media_geral)}</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption(f"Baseado em {len(medias)} loja(s) com dados")
                    except (ValueError, TypeError):
                        pass

        # ---- GRÁFICO ----
        with col_grafico:
            st.markdown("### 📈 Gráfico de Pontuação")

            modo_grafico = st.radio(
                "Tipo de gráfico",
                options=["Mensal", "Tendência Anual"],
                horizontal=True,
                key="modo_grafico"
            )

            fig = None
            if modo_grafico == "Mensal" and ano_sel:
                try:
                    ano = int(ano_sel)
                    fig = gerar_grafico_mensal(dados, estado_sel, loja_sel if not todas_lojas else "", ano, todas_lojas)
                except (ValueError, TypeError):
                    st.warning("Ano inválido.")
            elif modo_grafico == "Tendência Anual":
                fig = gerar_grafico_anual(dados, estado_sel, loja_sel if not todas_lojas else "", todas_lojas)

            if fig is not None:
                st.pyplot(fig)
            else:
                if estado_sel and ano_sel:
                    st.info("Sem dados numéricos para exibir. Insira pontuações nos campos mensais.")
                else:
                    st.info("Selecione filtros para visualizar o gráfico.")

    # ============================================================
    # TAB: TABELA DE DADOS
    # ============================================================
    with tab_tabela:
        st.markdown("### 📋 Tabela de Dados")

        # Filtrar dados conforme seleção
        filtrar_ano = False
        ano_filtro = None
        if ano_sel:
            try:
                ano_filtro = int(ano_sel)
                filtrar_ano = True
            except (ValueError, TypeError):
                filtrar_ano = False

        filtrados = []
        for r in dados:
            if estado_sel and str(r.get("estado", "")).strip().upper() != estado_sel.upper():
                continue
            if not todas_lojas and loja_sel and str(r.get("loja", "")).strip().upper() != loja_sel.upper():
                continue
            if filtrar_ano and r.get("ano") != ano_filtro:
                continue
            filtrados.append(r)

        if filtrados:
            # Preparar DataFrame
            rows = []
            for r in filtrados:
                row = {
                    "Estado": r.get("estado", ""),
                    "Loja": r.get("loja", ""),
                    "Ano": r.get("ano", ""),
                }
                for mk, mnome in MESES:
                    val = str(r.get(f"pont_{mk}", "")).strip()
                    if val and not eh_valor_texto(val):
                        try:
                            f = float(val.replace(",", "."))
                            row[mnome] = formatar_pontuacao(f)
                        except (ValueError, TypeError):
                            row[mnome] = val
                    elif val:
                        row[mnome] = val
                    else:
                        row[mnome] = ""

                media_val = safe_float(r.get("media", 0))
                row["MÉDIA"] = formatar_media(media_val) if media_val > 0 else ""
                row["Contrato"] = r.get("contrato", "")
                rows.append(row)

            if HAS_PANDAS:
                df = pd.DataFrame(rows)
                # Colorir MÉDIA baseado na pontuação
                def highlight_media(val):
                    try:
                        v = float(str(val).replace(",", "."))
                        if v >= 90:
                            return 'background-color: #dcfce7'
                        elif v >= 70:
                            return 'background-color: #fef9c3'
                        elif v > 0:
                            return 'background-color: #fee2e2'
                        return ''
                    except (ValueError, TypeError):
                        if val and val.strip():
                            return 'background-color: #f3f4f6'
                        return ''

                styled = df.style.map(highlight_media, subset=["MÉDIA"])
                st.dataframe(styled, use_container_width=True, hide_index=True, height=500)
            else:
                # Fallback sem pandas
                st.table(rows)

            st.caption(f"{len(filtrados)} registros")
        else:
            st.info("Nenhum registro encontrado para os filtros selecionados.")


if __name__ == "__main__":
    main()
